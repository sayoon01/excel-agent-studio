"""
File Manager — 전문적이고 깔끔한 파일 관리 인터페이스
"""
import streamlit as st
import pandas as pd
from pathlib import Path
from datetime import datetime
import sys, io, zipfile
from itertools import zip_longest

sys.path.insert(0, str(Path(__file__).parent.parent))
from utils.activity_log import add_entry

st.set_page_config(page_title="File Manager", page_icon="📂", layout="wide")

UPLOAD_DIR = Path("uploads")
UPLOAD_DIR.mkdir(exist_ok=True)

# ══════════════════════════════════════════════
# 캐시 유틸
# ══════════════════════════════════════════════

@st.cache_data(ttl=120, show_spinner=False)
def get_sheet_names(path_str: str) -> list[str]:
    fp = Path(path_str)
    if fp.suffix.lower() == ".csv":
        return ["Sheet1"]
    try:
        from openpyxl import load_workbook
        wb = load_workbook(fp, read_only=True, data_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    except Exception:
        return ["Sheet1"]


@st.cache_data(ttl=120, show_spinner=False)
def read_excel_smart(
    path_str: str,
    sheet_idx: int = 0,
    header_row: int = 0,
    n_header_rows: int = 1,
) -> pd.DataFrame:
    """
    병합 셀 해제 + 다중 헤더 행 지원 + 쉼표 포함 숫자 문자열 처리.

    n_header_rows=2 : 대제목(row1)+소제목(row2) 구조.
      소제목이 대제목과 다를 경우 소제목을 컬럼명으로 사용.
      예) 실행예산(row1) + 이월예산(row2) → 컬럼명 = 이월예산
    """
    fp = Path(path_str)
    if fp.suffix.lower() == ".csv":
        return pd.read_csv(fp)
    try:
        from openpyxl import load_workbook
        wb = load_workbook(fp, data_only=True)
        ws = wb.worksheets[sheet_idx] if sheet_idx < len(wb.worksheets) else wb.active

        # 병합 셀 해제: 좌상단 값으로 채우기
        for merged in list(ws.merged_cells.ranges):
            top = ws.cell(merged.min_row, merged.min_col).value
            ws.unmerge_cells(str(merged))
            for r in range(merged.min_row, merged.max_row + 1):
                for c in range(merged.min_col, merged.max_col + 1):
                    ws.cell(r, c).value = top

        data = [[cell.value for cell in row] for row in ws.iter_rows()]
        wb.close()

        if not data or header_row >= len(data):
            return pd.DataFrame()

        # ── 헤더 결합 ──
        if n_header_rows == 2 and header_row + 1 < len(data):
            row1 = data[header_row]
            row2 = data[header_row + 1]
            data_start = header_row + 2

            raw: list = []
            for h1, h2 in zip_longest(row1, row2, fillvalue=None):
                s1 = str(h1).strip() if h1 is not None else None
                s2 = str(h2).strip() if h2 is not None else None
                # 소제목이 대제목과 다른 경우(더 구체적) → 소제목 사용
                if s2 and s1 and s2 != s1:
                    raw.append(s2)
                elif s2 and not s1:
                    raw.append(s2)
                elif s1:
                    raw.append(s1)
                else:
                    raw.append(None)
        else:
            raw = data[header_row]
            data_start = header_row + 1

        # ── 중복 컬럼명 처리 ──
        seen: dict[str, int] = {}
        headers: list[str] = []
        for i, h in enumerate(raw):
            name = str(h).strip() if h is not None else f"col_{i}"
            if name in seen:
                seen[name] += 1
                name = f"{name}_{seen[name]}"
            else:
                seen[name] = 0
            headers.append(name)

        df = pd.DataFrame(data[data_start:], columns=headers)
        df = df.dropna(how="all").reset_index(drop=True)

        # 쉼표 포함 숫자 문자열('51,840,000') 포함 숫자 타입 추론
        def _coerce(s: pd.Series) -> pd.Series:
            cleaned = s.apply(lambda x: str(x).replace(",", "") if isinstance(x, str) else x)
            return pd.to_numeric(cleaned, errors="coerce")

        for col in df.columns:
            conv = _coerce(df[col])
            if conv.notna().sum() / max(len(df), 1) >= 0.7:
                df[col] = conv

        return df
    except Exception as e:
        try:
            df_fb = pd.read_excel(fp, sheet_name=sheet_idx, header=header_row)
            # 정수 컬럼명(0,1,2..) → col_N 변환
            if all(isinstance(c, int) for c in df_fb.columns):
                df_fb.columns = [f"col_{c}" for c in df_fb.columns]
            return df_fb
        except Exception:
            return pd.DataFrame({"오류": [str(e)]})


@st.cache_data(ttl=120, show_spinner=False)
def get_file_meta(path_str: str) -> dict:
    sheets = get_sheet_names(path_str)
    try:
        df = read_excel_smart(path_str, 0, 0)
        return {"sheets": sheets, "rows": len(df), "cols": len(df.columns)}
    except Exception:
        return {"sheets": sheets, "rows": 0, "cols": 0}


@st.cache_data(ttl=120, show_spinner=False)
def read_file_bytes(path_str: str) -> bytes:
    return Path(path_str).read_bytes()


def _excel_col_name(n: int) -> str:
    """0-based index → Excel 열 문자 (0→A, 25→Z, 26→AA …)"""
    result, n = "", n + 1
    while n > 0:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result


@st.cache_data(ttl=120, show_spinner=False)
def read_excel_raw(path_str: str, sheet_idx: int = 0) -> pd.DataFrame:
    """
    원본 보기 전용.
    - 헤더 없이 모든 행 포함
    - 열 이름 = Excel 열 문자 (A, B, C ...)
    - 행 인덱스 = Excel 행 번호 (1부터)
    - 병합 셀 → 좌상단 값으로 채워서 표시
    """
    fp = Path(path_str)
    if fp.suffix.lower() == ".csv":
        try:
            df = pd.read_csv(fp, header=None)
            df.columns = [_excel_col_name(i) for i in range(len(df.columns))]
            df.index   = range(1, len(df) + 1)
            return df
        except Exception as e:
            return pd.DataFrame({"오류": [str(e)]})
    try:
        from openpyxl import load_workbook
        wb = load_workbook(fp, data_only=True)
        ws = wb.worksheets[sheet_idx] if sheet_idx < len(wb.worksheets) else wb.active

        # 병합 셀 해제 후 좌상단 값으로 채우기
        for merged in list(ws.merged_cells.ranges):
            top_val = ws.cell(merged.min_row, merged.min_col).value
            ws.unmerge_cells(str(merged))
            for r in range(merged.min_row, merged.max_row + 1):
                for c in range(merged.min_col, merged.max_col + 1):
                    ws.cell(r, c).value = top_val

        data = [[cell.value for cell in row] for row in ws.iter_rows()]
        wb.close()

        if not data:
            return pd.DataFrame()

        n_cols = max(len(row) for row in data)
        padded = [row + [None] * (n_cols - len(row)) for row in data]
        cols   = [_excel_col_name(i) for i in range(n_cols)]

        df = pd.DataFrame(padded, columns=cols)
        df.index = range(1, len(df) + 1)
        return df
    except Exception as e:
        return pd.DataFrame({"오류": [str(e)]})


def fmt_raw(val) -> str:
    """원본 보기 포매터: 타입 변환 없이, None/NaN → 빈 문자열."""
    if val is None:
        return ""
    if isinstance(val, bool):
        return str(val)
    if isinstance(val, float):
        if pd.isna(val):
            return ""
        return f"{int(val):,}" if val == int(val) else f"{val:,.2f}"
    if isinstance(val, int):
        return f"{val:,}"
    from datetime import datetime as _dt
    if isinstance(val, _dt):
        return val.strftime("%Y-%m-%d %H:%M") if val.hour or val.minute else val.strftime("%Y-%m-%d")
    s = str(val)
    return "" if s in ("None", "nan", "NaN") else s


def fmt_val(val) -> str:
    if val is None:
        return ""
    if isinstance(val, float):
        if pd.isna(val):
            return ""
        return f"{int(val):,}" if val == int(val) else f"{val:,.2f}"
    if isinstance(val, int) and not isinstance(val, bool):
        return f"{val:,}"
    s = str(val)
    return "" if s in ("None", "nan", "NaN") else s


# ══════════════════════════════════════════════
# 다이얼로그
# ══════════════════════════════════════════════

@st.dialog("파일 업로드")
def upload_dialog():
    uploaded = st.file_uploader(
        "엑셀 또는 CSV 파일 선택",
        type=["xlsx", "xls", "csv"],
        accept_multiple_files=True,
        label_visibility="collapsed",
    )
    if uploaded:
        names = []
        for f in uploaded:
            (UPLOAD_DIR / f.name).write_bytes(f.getbuffer())
            names.append(f.name)
            add_entry("업로드", f.name, "uploaded")
        st.cache_data.clear()
        st.success(f"✅ {len(names)}개 업로드 완료")
        if st.button("닫기", type="primary", use_container_width=True):
            st.rerun()


@st.dialog("삭제 확인")
def confirm_delete_dialog(targets: list[str]):
    """targets: 삭제할 파일 경로 문자열 리스트"""
    names = [Path(s).name for s in targets]

    if len(names) == 1:
        st.warning(f"**{names[0]}** 을(를) 삭제합니다.\n\n이 작업은 되돌릴 수 없습니다.")
    else:
        st.warning(f"선택한 **{len(names)}개** 파일을 삭제합니다.\n\n이 작업은 되돌릴 수 없습니다.")
        with st.container():
            for n in names[:6]:
                st.markdown(f"- {n}")
            if len(names) > 6:
                st.caption(f"  外 {len(names) - 6}개 더...")

    st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)
    col_ok, col_cancel = st.columns(2)

    if col_ok.button("🗑 삭제", type="primary", use_container_width=True):
        for s in targets:
            fp = Path(s)
            if fp.exists():
                add_entry("삭제", fp.name, "deleted")
                fp.unlink()
        st.cache_data.clear()
        st.rerun()

    if col_cancel.button("취소", use_container_width=True):
        st.rerun()


@st.dialog("미리보기", width="large")
def preview_dialog(fp_str: str):
    fp = Path(fp_str)
    sheets = get_sheet_names(fp_str)
    n_sheets = len(sheets)

    # ── 시트 선택 ──
    if n_sheets > 1:
        sel  = st.selectbox("시트", sheets)
        sidx = sheets.index(sel)
    else:
        sidx = 0
        st.caption(f"시트: {sheets[0] if sheets else '—'}")

    cache_key = f"_ai_analysis_{fp.name}_{sidx}"

    # ── AI 자동 분석 (탭 진입 전 1회 실행, 이후 캐시 사용) ──
    from utils.ai_caller import call_ai_simple, load_active_config
    ai_cfg       = load_active_config()
    active_model = ai_cfg.get("active_model", "")

    if active_model and cache_key not in st.session_state:
        with st.spinner("AI가 데이터를 분석하는 중…"):
            df_ai    = read_excel_smart(fp_str, sidx, 0, 1)
            num_cols = df_ai.select_dtypes(include="number").columns.tolist()
            cat_cols = df_ai.select_dtypes(exclude="number").columns.tolist()

            col_sums = {
                c: df_ai[c].sum()
                for c in num_cols
                if pd.notna(df_ai[c].sum()) and df_ai[c].sum() != 0
            }

            plan_kw   = ["계획", "예산", "budget", "plan"]
            exec_kw   = ["실행", "집행", "지출", "executed", "actual"]
            plan_cols = [c for c in num_cols if any(k in c for k in plan_kw)]
            exec_cols = [c for c in num_cols if any(k in c for k in exec_kw)]

            ratios: dict[str, str] = {}
            for pc in plan_cols:
                for ec in exec_cols:
                    ps = df_ai[pc].sum()
                    es = df_ai[ec].sum()
                    if ps and ps > 0:
                        ratios[f"{pc} 대비 {ec} 집행률"] = (
                            f"{es / ps * 100:.1f}%  ({es:,.0f} / {ps:,.0f})"
                        )

            zero_cols = [c for c in num_cols if df_ai[c].sum() == 0]

            low_exec_rows: list[str] = []
            if plan_cols and exec_cols:
                pc0, ec0 = plan_cols[0], exec_cols[0]
                label_col = cat_cols[1] if len(cat_cols) > 1 else (cat_cols[0] if cat_cols else None)
                if label_col:
                    mask     = (df_ai[pc0] > 0) & (df_ai[ec0] / df_ai[pc0] < 0.3)
                    low_rows = df_ai[mask][label_col].dropna().unique().tolist()
                    low_exec_rows = [
                        str(v) for v in low_rows[:6]
                        if str(v) not in ("nan", "None", "합계", "소계")
                    ]

            cat_vals = {c: df_ai[c].dropna().unique().tolist()[:8] for c in cat_cols[:4]}

            sums_txt  = "\n".join(f"  {k}: {v:,.0f}" for k, v in list(col_sums.items())[:12])
            ratio_txt = "\n".join(f"  {k}: {v}" for k, v in ratios.items()) or "  (계획/실행 쌍 없음)"
            zero_txt  = ", ".join(zero_cols) or "없음"
            low_txt   = "\n".join(f"  · {v}" for v in low_exec_rows) or "  (없음 또는 파악 불가)"
            cat_txt   = "\n".join(f"  {c}: {vals}" for c, vals in cat_vals.items())
            sample_txt = df_ai.head(8).to_string(max_cols=10)

            prompt = f"""당신은 한국 공공기관·연구소 행정 담당자를 위한 Excel 분석 전문가입니다.
아래 데이터를 보고, 담당자가 5초 안에 핵심을 파악할 수 있도록 분석해주세요.

═══ 파일 정보 ═══
파일명: {fp.name}
크기: {len(df_ai):,}행 × {len(df_ai.columns)}열
컬럼 목록: {', '.join(df_ai.columns.tolist()[:20])}

═══ Python이 미리 계산한 수치 (이 숫자를 그대로 사용할 것) ═══
컬럼별 합계:
{sums_txt or "  (숫자 컬럼 없음)"}

집행률 (계획 대비 실행):
{ratio_txt}

전액 미집행 컬럼: {zero_txt}

집행률 30% 미만 항목:
{low_txt}

분류 항목 목록:
{cat_txt}

데이터 샘플 (합계·소계 행 포함):
{sample_txt}

═══ 출력 형식 (반드시 이 형식으로, 한국어만) ═══

## 이 표는 "[문서 유형 추론]" 입니다
(1~2줄. 어떤 업무 목적의 표인지)

## 핵심 수치
(위 계산값 기반으로만. pandas 통계 용어 금지. 집행률·비중·합계 등 실무 관점 3~5개)

## 주의할 항목
(집행률 낮은 항목, 0인 컬럼, 이상한 패턴 — 구체적 항목명 포함. 없으면 생략)

## AI Prompt에서 바로 쓸 수 있는 명령어
1. "(이 파일에 맞는 구체적 프롬프트 예시)"
2. "(이 파일에 맞는 구체적 프롬프트 예시)"
3. "(이 파일에 맞는 구체적 프롬프트 예시)"
"""
            st.session_state[cache_key] = call_ai_simple(prompt)

    # ── 탭 ──
    t_raw, t_clean, t_ai, t_log = st.tabs([
        "원본",
        "정리 데이터",
        "AI 분석",
        "작업 로그",
    ])

    # ══ 원본 탭 ══════════════════════════════════
    with t_raw:
        st.caption(
            "열 이름 = Excel 열 문자 (A, B, C…) · 행 번호 = Excel 행 번호 (1부터) · "
            "병합 셀은 좌상단 값으로 채워 표시"
        )
        with st.spinner(""):
            df_raw = read_excel_raw(fp_str, sidx)

        if df_raw.empty:
            st.warning("데이터를 읽을 수 없습니다.")
        else:
            st.dataframe(
                df_raw.map(fmt_raw),
                use_container_width=True,
                height=440,
            )
            st.caption(f"전체 {len(df_raw):,}행 × {len(df_raw.columns)}열")

    # ══ 정리 데이터 탭 ════════════════════════════
    with t_clean:
        c_hdr, c_nhdr = st.columns([1, 1.3])
        hdr = int(c_hdr.number_input(
            "헤더 시작 행", 0, 10, 0,
            help="0 = 첫째 행이 헤더 (0-based)\n원본 탭에서 행 번호를 확인 후 입력",
        ))
        n_hdr = c_nhdr.selectbox(
            "헤더 행 수", [1, 2], index=0,
            help="2 = 대제목+소제목 2행 구조\n예) 실행예산(1행) + 이월예산(2행) → 컬럼명 = 이월예산",
        )
        if hdr > 0 or n_hdr == 2:
            st.caption("소제목이 대제목과 다르면 소제목을 컬럼명으로 사용 · 중복 컬럼명은 _1, _2 추가")

        with st.spinner(""):
            df = read_excel_smart(fp_str, sidx, hdr, n_hdr)

        if df.empty:
            st.warning("데이터를 읽을 수 없습니다. 헤더 설정을 조정해보세요.")
        else:
            st.dataframe(df.head(200).map(fmt_val), use_container_width=True, height=300)
            if len(df) > 200:
                st.caption(f"처음 200행 표시 / 전체 {len(df):,}행")

            # 컬럼 통계
            with st.expander("컬럼 통계"):
                _num = df.select_dtypes(include="number").columns.tolist()
                _cat = df.select_dtypes(exclude="number").columns.tolist()
                if _num:
                    _stats = df[_num].agg(["count", "sum", "mean", "min", "max"])
                    _stats.index = ["건수", "합계", "평균", "최솟값", "최댓값"]
                    st.dataframe(
                        _stats.map(lambda x: f"{x:,.1f}" if pd.notna(x) else "—"),
                        use_container_width=True,
                    )
                for c in _cat[:6]:
                    uniq_vals = df[c].dropna().unique()
                    preview   = ", ".join(str(v) for v in uniq_vals[:5])
                    st.caption(
                        f"**{c}**: {len(uniq_vals)}개 고유값 — {preview}"
                        + ("…" if len(uniq_vals) > 5 else "")
                    )

            buf = io.BytesIO()
            df.to_excel(buf, index=False)
            st.download_button(
                f"⬇ {fp.stem}_정리본.xlsx 다운로드",
                buf.getvalue(),
                file_name=f"{fp.stem}_정리본.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    # ══ AI 분석 탭 ══════════════════════════════════
    with t_ai:
        if not active_model:
            st.warning(
                "모델이 연결되지 않았습니다. "
                "**Model Manager**에서 Ollama / OpenAI / Google 중 하나를 활성화하세요."
            )
        else:
            cached = st.session_state.get(cache_key)
            if cached:
                st.caption(f"분석 모델: **{active_model}**")
                st.markdown(cached)
                st.divider()
                col_re, _ = st.columns([1, 4])
                if col_re.button("↻ 다시 분석", use_container_width=True):
                    st.session_state.pop(cache_key, None)
                    st.rerun()
            else:
                st.info("분석 결과를 불러오지 못했습니다.")
                if st.button("분석 재시도", type="primary"):
                    st.rerun()

    # ══ 작업 로그 탭 ══════════════════════════════
    with t_log:
        from utils.activity_log import load_entries

        all_entries  = load_entries()
        file_entries = [e for e in all_entries if e.get("filename") == fp.name]

        # AI Prompt 코드 히스토리 중 이 파일을 언급한 항목
        code_hist = [
            h for h in st.session_state.get("code_history", [])
            if fp.name in h.get("prompt", "") or fp.name in str(h.get("files", []))
        ]

        if not file_entries and not code_hist:
            st.info("이 파일에 대한 작업 기록이 없습니다.")
        else:
            _LOG_STYLE = {
                "uploaded": ("#ede9fe", "#5b21b6", "업로드"),
                "processed": ("#d1fae5", "#065f46", "처리"),
                "analyzed":  ("#dbeafe", "#1e40af", "분석"),
                "executed":  ("#d1fae5", "#065f46", "실행"),
                "merged":    ("#d1fae5", "#065f46", "통합"),
                "saved":     ("#fef3c7", "#92400e", "저장"),
                "deleted":   ("#f3f4f6", "#374151", "삭제"),
                "error":     ("#fee2e2", "#991b1b", "오류"),
            }

            if file_entries:
                st.caption(f"총 {len(file_entries)}건의 기록")
                for entry in file_entries:
                    action = entry.get("action", "")
                    status = entry.get("status", "")
                    ts_str = entry.get("timestamp", "")
                    detail = entry.get("detail", "")
                    bg, fg, label = _LOG_STYLE.get(status, ("#f3f4f6", "#374151", status))

                    try:
                        age  = datetime.now() - datetime.fromisoformat(ts_str)
                        secs = age.total_seconds()
                        time_str = (
                            "방금"              if secs < 60    else
                            f"{int(secs//60)}분 전"    if secs < 3600  else
                            f"{int(secs//3600)}시간 전" if secs < 86400 else
                            f"{int(age.days)}일 전"
                        )
                    except Exception:
                        time_str = ts_str[:16] if ts_str else ""

                    st.markdown(
                        f'<div style="display:flex;align-items:center;gap:10px;'
                        f'padding:8px 12px;border:1px solid #e2e8f0;border-radius:8px;'
                        f'margin-bottom:5px;font-size:13px">'
                        f'<span style="background:{bg};color:{fg};padding:2px 8px;'
                        f'border-radius:12px;font-size:11px;font-weight:600">{label}</span>'
                        f'<span style="flex:1;color:#374151">{action}</span>'
                        f'<span style="color:#94a3b8;font-size:12px;white-space:nowrap">{time_str}</span>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )
                    if detail:
                        st.caption(f"  └ {detail}")

            if code_hist:
                if file_entries:
                    st.divider()
                st.markdown("**AI Prompt 처리 이력**")
                for h in code_hist[:10]:
                    prompt_preview = h.get("prompt", "")[:100]
                    st.markdown(
                        f'<div style="padding:7px 12px;border-left:3px solid #60a5fa;'
                        f'margin-bottom:5px;font-size:12px;color:#374151">'
                        f'{prompt_preview}{"…" if len(h.get("prompt",""))>100 else ""}'
                        f'</div>',
                        unsafe_allow_html=True,
                    )


# ══════════════════════════════════════════════
# CSS
# ══════════════════════════════════════════════

st.markdown("""
<style>
/* 파일 행 세로 중앙 정렬 */
[data-testid="stHorizontalBlock"] { align-items: center; }

/* 버튼 높이 통일 */
[data-testid="stDownloadButton"] > button,
[data-testid="stBaseButton-secondary"] {
    height: 32px !important;
    padding: 0 8px !important;
    font-size: 13px !important;
}

/* 삭제 버튼 — 빨간 계열 */
[data-testid="stColumn"]:has(.del-zone) button {
    color: #dc2626 !important;
    border-color: #fca5a5 !important;
}
[data-testid="stColumn"]:has(.del-zone) button:hover {
    background: #fef2f2 !important;
    border-color: #f87171 !important;
}

/* 구분선 */
.file-divider {
    border: none;
    border-top: 1px solid #f1f5f9;
    margin: 1px 0;
}
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════
# 헤더
# ══════════════════════════════════════════════

hl, hr = st.columns([5, 1])
with hl:
    st.markdown("## File Manager")
with hr:
    st.markdown("<div style='padding-top:8px'>", unsafe_allow_html=True)
    if st.button("⬆ Upload", type="primary", use_container_width=True):
        upload_dialog()
    st.markdown("</div>", unsafe_allow_html=True)

# ── 다이얼로그 트리거 ──
if "preview_trigger" in st.session_state:
    _tgt = st.session_state.pop("preview_trigger")
    preview_dialog(_tgt)

if "delete_trigger" in st.session_state:
    _del = st.session_state.pop("delete_trigger")
    confirm_delete_dialog(_del)


# ══════════════════════════════════════════════
# 검색 + 정렬
# ══════════════════════════════════════════════

sc, srt = st.columns([5, 1])
search  = sc.text_input("search", placeholder="파일명 검색…",
                         label_visibility="collapsed")
sort_by = srt.selectbox("sort", ["최신순", "이름순", "크기순"],
                         label_visibility="collapsed")

files = [f for f in UPLOAD_DIR.glob("*")
         if f.suffix.lower() in (".xlsx", ".xls", ".csv")]
if search:
    files = [f for f in files if search.lower() in f.name.lower()]

_sort = {
    "최신순": lambda f: f.stat().st_mtime,
    "이름순": lambda f: f.name.lower(),
    "크기순": lambda f: f.stat().st_size,
}
files.sort(key=_sort[sort_by], reverse=(sort_by != "이름순"))

if not files:
    st.markdown("<br>", unsafe_allow_html=True)
    st.info("업로드된 파일이 없습니다. 우측 상단 **Upload** 버튼을 눌러 파일을 추가하세요.")
    st.stop()

# ── 파일 수·용량 요약 ──
_total_bytes = sum(f.stat().st_size for f in files)
_total_str   = (f"{_total_bytes/1024:.0f} KB" if _total_bytes < 1_048_576
                else f"{_total_bytes/1_048_576:.1f} MB")
st.markdown(
    f'<div style="font-size:12px;color:#94a3b8;padding:4px 2px 6px">'
    f'{len(files)}개 파일 &nbsp;·&nbsp; {_total_str}</div>',
    unsafe_allow_html=True,
)

# ══════════════════════════════════════════════
# 파일 목록
# ══════════════════════════════════════════════

selected: list[Path] = []

# ── 컬럼 헤더 (파일 행과 동일 비율) ──
_COL = [0.25, 3.5, 1.8, 1.8, 0.38, 0.38, 0.3]
_HL  = "font-size:11px;color:#94a3b8;font-weight:600;letter-spacing:.04em"
hc, hn, hm, ht, *_ = st.columns(_COL)
hn.markdown(f'<span style="{_HL}">파일명</span>', unsafe_allow_html=True)
hm.markdown(f'<span style="{_HL}">크기 · 수정일</span>', unsafe_allow_html=True)
ht.markdown(f'<span style="{_HL}">시트 · 행·열</span>', unsafe_allow_html=True)
st.markdown(
    "<hr style='border:none;border-top:1px solid #e2e8f0;margin:2px 0 4px'>",
    unsafe_allow_html=True,
)

for i, fp in enumerate(files):
    stat     = fp.stat()
    size_str = (f"{stat.st_size/1024:.0f} KB"
                if stat.st_size < 1_048_576
                else f"{stat.st_size/1_048_576:.1f} MB")
    date_str = datetime.fromtimestamp(stat.st_mtime).strftime("%m/%d %H:%M")
    ext      = fp.suffix.lower()
    icon     = "📊" if ext in (".xlsx", ".xls") else "📋"

    meta      = get_file_meta(str(fp))
    sheets    = meta["sheets"]
    n_sheets  = len(sheets)
    sheet_lbl = f"{n_sheets}시트" if n_sheets > 1 else (sheets[0][:10] if sheets else "—")
    dim_lbl   = f"{meta['rows']:,}행 · {meta['cols']}열" if meta["rows"] else "—"

    # 체크 | 이름 | 크기·날짜 | 시트·행열 | 다운 | 미리보기 | 삭제
    cc, cn, cm, ct, cdl, cpv, cdel = st.columns(_COL)

    chk = cc.checkbox("", key=f"chk_{i}", label_visibility="collapsed")
    if chk:
        selected.append(fp)

    cn.markdown(
        f"<div style='display:flex;align-items:center;gap:8px'>"
        f"<span style='font-size:16px'>{icon}</span>"
        f"<span style='font-weight:500;font-size:13px'>{fp.name}</span>"
        f"</div>",
        unsafe_allow_html=True,
    )
    cm.markdown(
        f"<span style='font-size:12px;color:#94a3b8'>{size_str} · {date_str}</span>",
        unsafe_allow_html=True,
    )
    ct.markdown(
        f"<span style='font-size:12px;color:#94a3b8'>{sheet_lbl} · {dim_lbl}</span>",
        unsafe_allow_html=True,
    )

    cdl.download_button(
        "⬇", data=read_file_bytes(str(fp)),
        file_name=fp.name, key=f"dl_{i}",
        use_container_width=True, help="원본 파일 다운로드",
    )

    if cpv.button("◉", key=f"pv_{i}", use_container_width=True, help="미리보기"):
        st.session_state["preview_trigger"] = str(fp)
        st.rerun()

    # 삭제 버튼 — 빨간 스타일 마커
    cdel.markdown('<span class="del-zone" style="display:none"></span>', unsafe_allow_html=True)
    if cdel.button("🗑", key=f"del_{i}", use_container_width=True):
        st.session_state["delete_trigger"] = [str(fp)]
        st.rerun()

    st.markdown("<hr class='file-divider'>", unsafe_allow_html=True)


# ── 하단 안내 ──
st.markdown(
    '<div style="font-size:11px;color:#cbd5e1;text-align:center;padding:14px 0 4px">'
    '파일을 체크하면 AI Prompt 전송 · ZIP 다운로드 · 일괄 삭제 가능'
    '</div>',
    unsafe_allow_html=True,
)

# ══════════════════════════════════════════════
# 선택 파일 작업 바
# ══════════════════════════════════════════════

if not selected:
    st.stop()

st.markdown("<br>", unsafe_allow_html=True)
st.markdown(
    f'<div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:10px;'
    f'padding:10px 16px;display:flex;align-items:center;gap:10px">'
    f'<span style="font-weight:600;color:#1e40af;font-size:13px">✓ {len(selected)}개 선택됨</span>'
    f'<span style="color:#64748b;font-size:12px">'
    f'{" · ".join(f.name for f in selected[:3])}'
    f'{"  外…" if len(selected) > 3 else ""}</span>'
    f'</div>',
    unsafe_allow_html=True,
)
st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

a1, a2, a3 = st.columns(3)

with a1:
    if st.button("AI Prompt로 열기", type="primary", use_container_width=True):
        st.session_state["ai_prompt_preselect"] = [f.name for f in selected]
        st.switch_page("pages/2_AI_Prompt.py")

with a2:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for fp in selected:
            zf.write(fp, fp.name)
    buf.seek(0)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    st.download_button(
        "ZIP 다운로드",
        data=buf.read(),
        file_name=f"files_{ts}.zip",
        mime="application/zip",
        use_container_width=True,
    )

with a3:
    if st.button(f"🗑 선택 삭제 ({len(selected)}개)", use_container_width=True):
        st.session_state["delete_trigger"] = [str(fp) for fp in selected]
        st.rerun()
