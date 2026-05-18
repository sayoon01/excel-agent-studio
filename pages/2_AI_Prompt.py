"""
AI Prompt — 2-Phase 구조
  Phase 1 : LLM이 Action Plan(JSON) 생성
  Phase 2 : Python Tool이 실제 계산
  Phase 3 : LLM이 결과만 설명
  Fallback : 계획 실패 시 코드 생성·실행 방식
"""
import streamlit as st
import pandas as pd
from pathlib import Path
from datetime import datetime
import json, io, sys, traceback, textwrap, re, base64
import matplotlib
matplotlib.use("Agg")
matplotlib.rcParams["axes.unicode_minus"] = False
import matplotlib.pyplot as plt

sys.path.insert(0, str(Path(__file__).parent.parent))
from utils.activity_log import add_entry
from utils.task_router import (
    RouteResult,
    call_with_route,
    classify_task,
    has_model_pool,
    route_for_prompt,
    route_for_role,
    stream_with_route,
    TASK_LABELS,
)
from utils.styles import apply_global_css

st.set_page_config(page_title="AI Prompt", page_icon="💬", layout="wide")
apply_global_css()

UPLOAD_DIR   = Path("uploads")
OUTPUT_DIR   = Path("outputs")
HISTORY_FILE = Path("code_history.json")
CONFIG_FILE  = Path("model_config.json")
UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

for key, default in [("messages", []), ("tokens_used", 0)]:
    if key not in st.session_state:
        st.session_state[key] = default


# ══════════════════════════════════════════════
# 히스토리
# ══════════════════════════════════════════════

def load_history() -> list:
    if HISTORY_FILE.exists():
        try:
            return json.loads(HISTORY_FILE.read_text(encoding="utf-8"))
        except Exception:
            return []
    return []

def save_to_history(prompt: str, code: str) -> None:
    hist = load_history()
    for h in hist:
        if h.get("code", "").strip() == code.strip():
            h["prompt"] = prompt; h["timestamp"] = datetime.now().isoformat()
            HISTORY_FILE.write_text(json.dumps(hist[:50], ensure_ascii=False, indent=2)); return
    hist.insert(0, {"prompt": prompt[:80], "code": code, "timestamp": datetime.now().isoformat()})
    HISTORY_FILE.write_text(json.dumps(hist[:50], ensure_ascii=False, indent=2))


# ══════════════════════════════════════════════
# 모델 설정 + 사이드바
# ══════════════════════════════════════════════

def load_model_config() -> dict:
    if CONFIG_FILE.exists():
        try:
            return json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {"openai": {}, "ollama": {"host": "http://localhost:11434"}, "google": {}}

cfg = load_model_config()

with st.sidebar:
    st.markdown("### 자동 모델")
    if not has_model_pool(cfg):
        st.warning("모델 풀 미등록 — Model Manager에서 Ollama/API를 연결하세요")
    _route = st.session_state.get("auto_route")
    if _route:
        st.markdown(
            f'<div style="background:#d1fae5;color:#065f46;padding:4px 10px;'
            f'border-radius:8px;font-size:12px;font-weight:600;margin-bottom:4px">'
            f'● {_route.display}</div>',
            unsafe_allow_html=True,
        )
        st.caption(f"{_route.task_label} · 자동 선택됨")
    else:
        st.caption("질문을 입력하면 작업 유형에 맞는 모델이 자동 선택됩니다")

    st.divider()
    st.markdown("### 파일 선택")
    excel_files = sorted([f for f in UPLOAD_DIR.glob("*")
                          if f.suffix.lower() in (".xlsx", ".xls", ".csv")], key=lambda f: f.name)
    if excel_files:
        all_names   = [f.name for f in excel_files]
        preselect   = st.session_state.pop("ai_prompt_preselect", None)
        default_sel = ([n for n in preselect if n in all_names] if preselect else all_names)
        if preselect:
            st.info(f"File Manager에서 {len(default_sel)}개 파일이 선택됐습니다.")
        selected_files = st.multiselect("Select files", all_names, default=default_sel)
        st.caption(f"✅ {len(selected_files)}개 파일 선택됨")
    else:
        selected_files = []
        st.info("File Manager에서 파일을 먼저 업로드하세요")

    st.divider()
    history = load_history()
    if history:
        st.markdown("### 코드 히스토리")
        st.caption("클릭하면 해당 프롬프트를 재실행합니다")
        for i, h in enumerate(history[:8]):
            if st.button(f"↩ {h.get('prompt', f'작업 {i+1}')[:45]}", key=f"hist_{i}", use_container_width=True):
                st.session_state["rerun_prompt"] = h.get("prompt", "")
                st.session_state["rerun_code"]   = h.get("code", "")
                st.rerun()
        if st.button("히스토리 초기화", use_container_width=True):
            HISTORY_FILE.unlink(missing_ok=True); st.rerun()
        st.divider()

    has_msgs = bool(st.session_state.get("messages"))
    if has_msgs:
        msgs = st.session_state.get("messages", [])
        ts   = __import__("datetime").datetime.now().strftime("%Y%m%d_%H%M%S")
        md_lines = [f"# AI Chat — {ts}\n"]
        for _m in msgs:
            _role = "User" if _m["role"] == "user" else "AI"
            md_lines.append(f"### {_role}\n\n{_m.get('content','')}\n\n---\n")
        md_content = "\n".join(md_lines)
        st.download_button(
            "대화 저장 (MD)", data=md_content.encode("utf-8"),
            file_name=f"chat_{ts}.md", mime="text/markdown",
            use_container_width=True,
        )
    if st.button("대화 초기화", use_container_width=True):
        st.session_state.messages = []; st.rerun()


# ══════════════════════════════════════════════
# 헬퍼
# ══════════════════════════════════════════════

def _excel_col_name(n: int) -> str:
    result, n = "", n + 1
    while n > 0:
        n, r = divmod(n - 1, 26); result = chr(65 + r) + result
    return result

def _col_letter_to_idx(s: str) -> int:
    result = 0
    for c in s.upper():
        result = result * 26 + (ord(c) - 64)
    return result - 1

def find_column(df: pd.DataFrame, hint: str | None) -> str | None:
    if hint is None or df is None: return None
    if hint in df.columns: return hint
    for col in df.columns:
        if col.lower() == hint.lower(): return col
    if len(hint) <= 2 and hint.replace("A","").replace("Z","").isalpha():
        idx = _col_letter_to_idx(hint)
        if 0 <= idx < len(df.columns): return df.columns[idx]
    for col in df.columns:
        if hint.lower() in col.lower(): return col
    return None

def _get_df(ref: str | None, ctx: dict, dfs: dict) -> pd.DataFrame | None:
    if not ref:
        for v in reversed(list(ctx.values())):
            if isinstance(v, dict) and isinstance(v.get("dataframe"), pd.DataFrame):
                return v["dataframe"]
        return next(iter(dfs.values()), None)
    if ref in ctx:
        v = ctx[ref]
        if isinstance(v, pd.DataFrame): return v
        if isinstance(v, dict) and isinstance(v.get("dataframe"), pd.DataFrame): return v["dataframe"]
    if ref in dfs: return dfs[ref]
    for fname, df in dfs.items():
        if ref.lower() in fname.lower() or Path(fname).stem.lower() == ref.lower(): return df
    return next(iter(dfs.values()), None)


# ══════════════════════════════════════════════
# Python Tools  (각 Tool → structured dict 반환)
# ══════════════════════════════════════════════

def tool_inspect_column(args: dict, ctx: dict, dfs: dict) -> dict:
    df  = _get_df(args.get("file"), ctx, dfs)
    col = find_column(df, args.get("column")) if df is not None else None
    if df is None or col is None:
        return {"text": f"컬럼 '{args.get('column')}' 을 찾을 수 없습니다."}
    ecol   = _excel_col_name(df.columns.get_loc(col))
    series = df[col]
    vals   = {f"{ecol}{i+2}": v for i, v in enumerate(series) if pd.notna(v)}
    result = {
        "file": args.get("file", list(dfs.keys())[0] if dfs else ""),
        "column": col, "excel_col": ecol,
        "dtype": "숫자형" if pd.api.types.is_numeric_dtype(series) else "텍스트형",
        "total_rows": len(series), "null_count": int(series.isnull().sum()),
        "unique_count": int(series.nunique()),
        "values": dict(list(vals.items())[:30]),
    }
    if pd.api.types.is_numeric_dtype(series):
        result["stats"] = {"합계": series.sum(), "평균": round(series.mean(), 2),
                           "최솟값": series.min(), "최댓값": series.max()}
    return result


def tool_compare_files(args: dict, ctx: dict, dfs: dict) -> dict:
    keys = list(dfs.keys())
    fa   = args.get("file_a") or (keys[0] if keys else None)
    fb   = args.get("file_b") or (keys[1] if len(keys) > 1 else None)
    if not fa or not fb:
        return {"text": "비교할 파일이 2개 이상 필요합니다."}
    dfa = _get_df(fa, ctx, dfs)
    dfb = _get_df(fb, ctx, dfs)
    if dfa is None or dfb is None:
        return {"text": "파일 로드 실패"}
    key_col = args.get("key_col")
    kc = find_column(dfa, key_col) if key_col else None
    result = {
        "file_a": fa, "file_b": fb,
        "shape": {"a": list(dfa.shape), "b": list(dfb.shape)},
        "columns_only_a": list(set(dfa.columns) - set(dfb.columns)),
        "columns_only_b": list(set(dfb.columns) - set(dfa.columns)),
    }
    if kc and kc in dfa.columns and kc in dfb.columns:
        merged = dfa.merge(dfb, on=kc, how="outer", suffixes=("_이전", "_이후"), indicator=True)
        result["added"]   = merged[merged["_merge"] == "right_only"][kc].tolist()[:20]
        result["removed"] = merged[merged["_merge"] == "left_only"][kc].tolist()[:20]
        common_cols = [c for c in dfa.columns if c != kc and c in dfb.columns
                       and pd.api.types.is_numeric_dtype(dfa[c])]
        changes = []
        for _, row in merged[merged["_merge"] == "both"].iterrows():
            for col in common_cols:
                va, vb = row.get(f"{col}_이전"), row.get(f"{col}_이후")
                if pd.notna(va) and pd.notna(vb) and va != vb:
                    changes.append({"항목": row[kc], "컬럼": col,
                                    "이전": va, "이후": vb, "증감": round(vb - va, 2)})
        result["changes"]   = changes[:20]
        result["dataframe"] = merged.drop(columns=["_merge"])
    else:
        # Shape + column 수준 비교만
        num_a = dfa.select_dtypes(include="number")
        num_b = dfb.select_dtypes(include="number")
        common = [c for c in num_a.columns if c in num_b.columns]
        result["sum_diff"] = {c: {"이전": num_a[c].sum(), "이후": num_b[c].sum(),
                                  "증감": round(num_b[c].sum() - num_a[c].sum(), 2)}
                              for c in common[:10]}
    return result


def tool_clean_table(args: dict, ctx: dict, dfs: dict) -> dict:
    df = _get_df(args.get("file"), ctx, dfs)
    if df is None: return {"text": "데이터를 찾을 수 없습니다."}
    df = df.copy(); orig = len(df)
    if args.get("remove_subtotals", True):
        kws  = ["합계", "소계", "계", "총계", "합", "total", "sum", "subtotal"]
        mask = pd.Series([False] * len(df), index=df.index)
        for col in df.select_dtypes(exclude="number").columns:
            mask |= df[col].astype(str).str.strip().str.lower().isin(kws)
        df = df[~mask].reset_index(drop=True)
    df = df.dropna(how="all").reset_index(drop=True)
    return {"removed_rows": orig - len(df), "cleaned_rows": len(df),
            "original_rows": orig, "dataframe": df, "output_name": "clean_table"}


def tool_aggregate(args: dict, ctx: dict, dfs: dict) -> dict:
    df = _get_df(args.get("file"), ctx, dfs)
    if df is None: return {"text": "데이터를 찾을 수 없습니다."}
    # 소계/합계 행 자동 제거 — 이중 집계 방지
    _sub_kws = {"합계", "소계", "계", "총계", "합", "total", "sum", "subtotal"}
    _sub_mask = pd.Series(False, index=df.index)
    for _c in df.select_dtypes(exclude="number").columns:
        _sub_mask |= df[_c].astype(str).str.strip().str.lower().isin(_sub_kws)
    if _sub_mask.any():
        df = df[~_sub_mask].reset_index(drop=True)
    gc  = find_column(df, args.get("group_by"))
    vc  = find_column(df, args.get("value_col"))
    fn  = args.get("func", "sum")
    if gc is None:
        result_df = df.select_dtypes(include="number").agg(fn).to_frame(name=fn).T
    elif vc is None:
        result_df = df.groupby(gc).agg(fn, numeric_only=True).reset_index()
    else:
        result_df = df.groupby(gc)[vc].agg(fn).reset_index()
    # 파일별 원본값 샘플 — 여러 파일 로드 시 근거 데이터로 활용
    per_file_sample: dict = {}
    if len(dfs) > 1 and gc and vc and gc in result_df.columns and vc in result_df.columns:
        top_groups = result_df[gc].head(5).tolist()
        for fname_f, df_f in dfs.items():
            if gc not in df_f.columns or vc not in df_f.columns:
                continue
            for grp in top_groups:
                match = df_f[df_f[gc] == grp][vc].dropna()
                if not match.empty:
                    per_file_sample.setdefault(str(grp), {})[fname_f] = float(match.iloc[0])

    return {"group_by": gc, "value_col": vc, "func": fn,
            "result": result_df.head(20).to_dict("records"), "dataframe": result_df,
            "per_file_sample": per_file_sample}


def tool_filter_rows(args: dict, ctx: dict, dfs: dict) -> dict:
    df  = _get_df(args.get("file"), ctx, dfs)
    if df is None: return {"text": "데이터를 찾을 수 없습니다."}
    col = find_column(df, args.get("column"))
    if col is None: return {"text": f"컬럼 '{args.get('column')}' 을 찾을 수 없습니다."}
    cond, val = args.get("condition", "notnull"), args.get("value")
    try:
        if cond == "eq":       mask = df[col] == val
        elif cond == "neq":    mask = df[col] != val
        elif cond == "gt":     mask = pd.to_numeric(df[col], errors="coerce") > float(val)
        elif cond == "lt":     mask = pd.to_numeric(df[col], errors="coerce") < float(val)
        elif cond == "gte":    mask = pd.to_numeric(df[col], errors="coerce") >= float(val)
        elif cond == "lte":    mask = pd.to_numeric(df[col], errors="coerce") <= float(val)
        elif cond == "contains": mask = df[col].astype(str).str.contains(str(val), na=False)
        elif cond == "isnull": mask = df[col].isnull()
        else:                  mask = df[col].notnull()
    except Exception:
        mask = pd.Series([True] * len(df))
    result_df = df[mask].reset_index(drop=True)
    result_df.index = range(2, len(result_df) + 2)
    return {"column": col, "condition": cond, "value": val,
            "matched": len(result_df), "total": len(df),
            "sample": result_df.head(10).to_dict("records"), "dataframe": result_df}


def tool_calculate_ratio(args: dict, ctx: dict, dfs: dict) -> dict:
    df = _get_df(args.get("file"), ctx, dfs)
    if df is None: return {"text": "데이터를 찾을 수 없습니다."}
    plan_kw = ["계획", "예산", "budget", "plan"]
    exec_kw = ["실행", "집행", "지출", "executed", "actual"]
    num_cols = df.select_dtypes(include="number").columns.tolist()
    pc = find_column(df, args.get("plan_col")) or next((c for c in num_cols if any(k in c for k in plan_kw)), None)
    ec = find_column(df, args.get("exec_col")) or next((c for c in num_cols if any(k in c for k in exec_kw)), None)
    if not pc or not ec:
        return {"text": f"계획/실행 컬럼을 찾을 수 없습니다. 현재 숫자 컬럼: {num_cols}"}
    lc = find_column(df, args.get("label_col")) or (
        df.select_dtypes(exclude="number").columns[0]
        if len(df.select_dtypes(exclude="number").columns) else None)
    result_df = df.copy()
    result_df["집행률(%)"] = (result_df[ec] / result_df[pc].replace(0, float("nan")) * 100).round(1)
    total_p, total_e = df[pc].sum(), df[ec].sum()
    rows = []
    if lc:
        for i, row in result_df.iterrows():
            pv, ev = row.get(pc, 0), row.get(ec, 0)
            if pd.notna(pv) and pv > 0:
                rows.append({"항목": row.get(lc, f"행{i+2}"),
                             "계획": pv, "실행": ev,
                             "집행률(%)": round(ev / pv * 100, 1)})
    return {"plan_col": pc, "exec_col": ec, "label_col": lc,
            "total_plan": total_p, "total_exec": total_e,
            "total_ratio": round(total_e / total_p * 100, 1) if total_p else 0,
            "rows": rows[:25], "dataframe": result_df}


def tool_fill_missing(args: dict, ctx: dict, dfs: dict) -> dict:
    df = _get_df(args.get("file"), ctx, dfs)
    if df is None: return {"text": "데이터를 찾을 수 없습니다."}
    df = df.copy()
    method  = args.get("method", "zero")
    cols    = [find_column(df, c) for c in args.get("columns", [])] if args.get("columns") else list(df.columns)
    cols    = [c for c in cols if c]
    before  = int(df.isnull().sum().sum())
    for col in cols:
        if col not in df.columns: continue
        if method == "zero":   df[col] = df[col].fillna(0)
        elif method == "ffill": df[col] = df[col].ffill()
        elif method == "bfill": df[col] = df[col].bfill()
        elif method == "mean" and pd.api.types.is_numeric_dtype(df[col]):
            df[col] = df[col].fillna(df[col].mean())
    return {"method": method, "before_null": before,
            "after_null": int(df.isnull().sum().sum()),
            "filled": before - int(df.isnull().sum().sum()), "dataframe": df}


def tool_visualize(args: dict, ctx: dict, dfs: dict) -> dict:
    df = _get_df(args.get("file") or args.get("source_step"), ctx, dfs)
    if df is None: return {"text": "차트 생성 실패: 데이터 없음"}
    num_cols = df.select_dtypes(include="number").columns.tolist()
    cat_cols = df.select_dtypes(exclude="number").columns.tolist()
    x = find_column(df, args.get("x")) or (cat_cols[0] if cat_cols else df.columns[0])
    y = find_column(df, args.get("y")) or (num_cols[0] if num_cols else df.columns[-1])
    chart_type = args.get("chart_type", "bar")
    plot_df = df.dropna(subset=[x, y]).head(20)
    plt.close("all")
    fig, ax = plt.subplots(figsize=(10, 4))
    if chart_type == "bar":
        ax.bar(range(len(plot_df)), plot_df[y], color="#60a5fa", edgecolor="white")
        ax.set_xticks(range(len(plot_df)))
        ax.set_xticklabels(plot_df[x].astype(str), rotation=35, ha="right", fontsize=8)
    elif chart_type == "line":
        ax.plot(range(len(plot_df)), plot_df[y], marker="o", color="#60a5fa", linewidth=2)
        ax.set_xticks(range(len(plot_df)))
        ax.set_xticklabels(plot_df[x].astype(str), rotation=35, ha="right", fontsize=8)
    elif chart_type == "pie":
        ax.pie(plot_df[y].abs(), labels=plot_df[x].astype(str), autopct="%1.1f%%", startangle=90)
    ax.set_title(str(y), fontsize=11)
    ax.spines[["top", "right"]].set_visible(False)
    plt.tight_layout()
    buf = io.BytesIO(); fig.savefig(buf, format="png", dpi=120, bbox_inches="tight"); buf.seek(0)
    chart_b64 = base64.b64encode(buf.read()).decode(); plt.close("all")
    return {"chart_b64": chart_b64, "x": x, "y": y, "rows": len(plot_df)}


def tool_export(args: dict, ctx: dict, dfs: dict) -> dict:
    src = args.get("source_step") or args.get("file")
    df  = _get_df(src, ctx, dfs)
    if df is None: return {"text": "저장할 데이터가 없습니다."}
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    name = args.get("filename", f"result_{ts}.xlsx")
    if not name.endswith((".xlsx", ".csv")): name += ".xlsx"
    out  = OUTPUT_DIR / name
    if name.endswith(".csv"): df.to_csv(out, index=False, encoding="utf-8-sig")
    else: df.to_excel(out, index=False)
    return {"output_file": str(out), "filename": name, "rows": len(df), "cols": len(df.columns)}


def tool_list_files(args: dict, ctx: dict, dfs: dict) -> dict:
    """업로드된 파일 목록과 상태 반환."""
    files = sorted(
        [f for f in UPLOAD_DIR.glob("*")
         if f.suffix.lower() in {".xlsx", ".xls", ".csv"} and f.name != ".gitkeep"],
        key=lambda f: f.stat().st_mtime, reverse=True,
    )
    rows = []
    for f in files:
        try:
            df_tmp = pd.read_csv(f) if f.suffix.lower() == ".csv" else _read_excel_smart(f)
            audit  = _audit_file(f.name, df_tmp)
            status = "⚠ " + " / ".join(audit["warnings"]) if audit["warnings"] else "✔ 정상"
            rows.append({
                "파일명": f.name,
                "전체 행": audit["rows"],
                "전체 열": audit["cols"],
                "텍스트 열": audit["text_col_count"],
                "숫자 열": audit["num_col_count"],
                "데이터 있는 행": audit["filled_rows"],
                "텍스트 입력 행": audit["text_filled_rows"],
                "크기": f"{f.stat().st_size / 1024:.0f} KB",
                "상태": status,
            })
        except Exception as e:
            rows.append({
                "파일명": f.name, "전체 행": "?", "전체 열": "?",
                "텍스트 열": "?", "숫자 열": "?",
                "데이터 있는 행": "?", "텍스트 입력 행": "?",
                "크기": "?", "상태": f"오류: {e}",
            })
    result_df = pd.DataFrame(rows) if rows else pd.DataFrame()
    return {"text": f"업로드된 파일 {len(rows)}개", "dataframe": result_df}


def tool_preview_file(args: dict, ctx: dict, dfs: dict) -> dict:
    """파일 데이터 미리보기 (처음 N행)."""
    fname  = args.get("file") or next(iter(dfs), "")
    n_rows = int(args.get("rows", 20))
    if fname not in dfs:
        fp = UPLOAD_DIR / fname
        if fp.exists():
            try:
                dfs[fname] = pd.read_csv(fp) if fp.suffix.lower() == ".csv" else _read_excel_smart(fp)
            except Exception as e:
                return {"text": f"'{fname}' 읽기 실패: {e}"}
        else:
            return {"text": f"파일 '{fname}'을 찾을 수 없습니다."}
    df = dfs[fname]
    return {
        "text": f"{fname} — {len(df):,}행 × {len(df.columns)}열 (처음 {min(n_rows, len(df))}행 표시)",
        "dataframe": df.head(n_rows),
    }


# ── Tool Registry ──
TOOL_REGISTRY: dict = {
    "list_files":       tool_list_files,
    "preview_file":     tool_preview_file,
    "inspect_column":   tool_inspect_column,
    "compare_files":    tool_compare_files,
    "clean_table":      tool_clean_table,
    "aggregate":        tool_aggregate,
    "filter_rows":      tool_filter_rows,
    "calculate_ratio":  tool_calculate_ratio,
    "fill_missing":     tool_fill_missing,
    "visualize":        tool_visualize,
    "export":           tool_export,
}

TOOL_ROLES: dict[str, str] = {
    "list_files":       "analysis",
    "preview_file":     "analysis",
    "visualize":        "code",
    "inspect_column":   "analysis",
    "compare_files":    "analysis",
    "clean_table":      "analysis",
    "aggregate":        "analysis",
    "filter_rows":      "analysis",
    "calculate_ratio":  "analysis",
    "fill_missing":     "analysis",
    "export":           "analysis",
}

TOOL_DESCRIPTIONS = """
- list_files       : 업로드 파일 목록·상태 표시  (args: 없음) ← "파일 목록 보여줘" 등
- preview_file     : 파일 데이터 미리보기        (args: file, rows=20) ← "데이터 보여줘" 등
- inspect_column   : 특정 열 상세 분석           (args: file, column)
- compare_files    : 두 파일 차이 비교           (args: file_a, file_b, key_col)
- clean_table      : 소계/빈행 제거              (args: file, remove_subtotals)
- aggregate        : 그룹별 집계                 (args: file, group_by, value_col, func)
- filter_rows      : 조건으로 행 필터            (args: file, column, condition[eq/gt/lt/contains/isnull], value)
- calculate_ratio  : 집행률/비율 계산            (args: file, plan_col, exec_col, label_col)
- fill_missing     : 빈 셀 채우기               (args: file, method[zero/ffill/mean], columns)
- visualize        : 차트 생성                   (args: file, chart_type[bar/line/pie], x, y)
- export           : 결과 엑셀/CSV 저장         (args: source_step, filename)
""".strip()

# ── 추천 타입 메타 (icon, bg, fg) ──
_TYPE_META: dict[str, tuple] = {
    "analyze": ("📊", "#eff6ff", "#1e40af"),
    "clean":   ("🧹", "#fffbeb", "#92400e"),
    "save":    ("📁", "#f0fdf4", "#065f46"),
    "insight": ("🧠", "#faf5ff", "#5b21b6"),
}

# ── Workflow 기반 정적 추천 (LLM 실패 시 fallback) ──
_WORKFLOW_FALLBACK: dict[str, list[dict]] = {
    "compare_files": [
        {"type": "analyze", "text": "증가 항목만 필터링해줘"},
        {"type": "analyze", "text": "감소 항목만 필터링해줘"},
        {"type": "analyze", "text": "신규 추가된 항목 추출해줘"},
        {"type": "save",    "text": "변경 결과 엑셀로 저장해줘"},
        {"type": "insight", "text": "가장 많이 변한 항목 설명해줘"},
    ],
    "aggregate": [
        {"type": "analyze", "text": "집계 결과를 막대 차트로 보여줘"},
        {"type": "analyze", "text": "상위 5개 항목만 보여줘"},
        {"type": "save",    "text": "집계 결과 엑셀로 저장해줘"},
        {"type": "insight", "text": "이상값 있는 항목 설명해줘"},
    ],
    "calculate_ratio": [
        {"type": "analyze", "text": "집행률 30% 미만 항목 필터링해줘"},
        {"type": "analyze", "text": "집행률을 막대 차트로 시각화해줘"},
        {"type": "analyze", "text": "집행률 상위·하위 5개 비교해줘"},
        {"type": "save",    "text": "집행률 분석 결과 엑셀로 저장해줘"},
        {"type": "insight", "text": "저조한 집행 항목 원인 요약해줘"},
    ],
    "inspect_column": [
        {"type": "analyze", "text": "이 열 기준으로 그룹별 합계 계산해줘"},
        {"type": "analyze", "text": "이 열을 막대 차트로 시각화해줘"},
        {"type": "clean",   "text": "빈 값을 0으로 채워줘"},
        {"type": "insight", "text": "이상값(outlier) 탐지해줘"},
    ],
    "filter_rows": [
        {"type": "save",    "text": "필터 결과 엑셀로 저장해줘"},
        {"type": "analyze", "text": "필터 결과를 차트로 시각화해줘"},
        {"type": "analyze", "text": "집행률 계산해줘"},
        {"type": "insight", "text": "필터된 항목 특징 요약해줘"},
    ],
    "visualize": [
        {"type": "analyze", "text": "상위 10개만 차트로 보여줘"},
        {"type": "analyze", "text": "파이 차트로 변경해줘"},
        {"type": "save",    "text": "차트 데이터 엑셀로 저장해줘"},
        {"type": "insight", "text": "차트에서 이상 패턴 설명해줘"},
    ],
    "clean_table": [
        {"type": "analyze", "text": "정리된 데이터로 집행률 계산해줘"},
        {"type": "analyze", "text": "컬럼별 합계 보여줘"},
        {"type": "save",    "text": "정리된 데이터 엑셀로 저장해줘"},
        {"type": "insight", "text": "정리 전후 차이 요약해줘"},
    ],
    "fill_missing": [
        {"type": "analyze", "text": "컬럼별 합계 다시 계산해줘"},
        {"type": "analyze", "text": "집행률 계산해줘"},
        {"type": "save",    "text": "처리된 데이터 저장해줘"},
        {"type": "insight", "text": "결측치 패턴 분석해줘"},
    ],
    "export": [
        {"type": "analyze", "text": "저장한 파일 내용 요약해줘"},
        {"type": "insight", "text": "전체 분석 결과 요약 리포트 만들어줘"},
    ],
}


def _summarize_tool_result(tool: str, res: dict) -> str:
    """Tool 결과를 LLM 프롬프트용 한 줄 요약으로 변환."""
    if tool == "compare_files":
        return (f"파일 비교 — 추가 {len(res.get('added',[]))}행·"
                f"삭제 {len(res.get('removed',[]))}행·"
                f"변경 {len(res.get('changes',[]))}건"
                + (" + 차트" if "chart_b64" in res else ""))
    if tool == "aggregate":
        return f"그룹별 집계 ({res.get('func','sum')}) by {res.get('group_by','')}"
    if tool == "calculate_ratio":
        return f"집행률 계산 — 전체 {res.get('total_ratio',0)}%"
    if tool == "inspect_column":
        stats = res.get("stats", {})
        return (f"컬럼 분석 — {res.get('column','')} ({res.get('dtype','')})"
                + (f", 합계={stats.get('합계','')}" if stats else ""))
    if tool == "filter_rows":
        return f"필터 — {res.get('total',0)}행 중 {res.get('matched',0)}행 해당"
    if tool == "visualize":
        return f"차트 생성 — {res.get('y','')} vs {res.get('x','')} ({res.get('rows',0)}행)"
    if tool == "clean_table":
        return f"테이블 정리 — {res.get('removed_rows',0)}행 제거, {res.get('cleaned_rows',0)}행 남음"
    if tool == "fill_missing":
        return f"결측치 처리 — {res.get('filled',0)}개 채움 ({res.get('method','')})"
    if tool == "export":
        return f"파일 저장 — {res.get('filename','')}"
    return tool


def generate_suggestions_ai(prompt: str, plan: dict, tool_results: list) -> list[dict]:
    """
    실제 실행 결과 기반으로 LLM이 다음 workflow 단계를 추천.
    LLM 호출 실패 시 _WORKFLOW_FALLBACK 정적 추천으로 대체.
    """
    ctx_lines  = [_summarize_tool_result(r["tool"], r["result"]) for r in tool_results]
    has_chart  = any("chart_b64" in r["result"] for r in tool_results)
    has_df     = any(isinstance(r["result"].get("dataframe"), pd.DataFrame) for r in tool_results)

    sys_prompt = """너는 데이터 분석 워크플로우 전문가야.
방금 완료된 작업 결과를 보고, 사용자가 다음에 할 만한 실용적인 업무 작업 4개를 추천해.

규칙:
1. "방금 한 작업과 비슷한 질문" 절대 금지 — 반드시 다음 단계 작업
2. type: analyze(분석·시각화), clean(정제), save(저장·보고서), insight(AI 설명) 중 하나
3. 구체적이고 즉시 실행 가능한 작업 (예: "증가 항목만 필터링해줘")
4. JSON 배열만 출력 (설명 없이)

출력:
[{"type":"analyze","text":"..."},{"type":"save","text":"..."},...]"""

    user_msg = (f"사용자 질문: {prompt}\n\n"
                f"완료된 작업:\n" + "\n".join(f"- {l}" for l in ctx_lines) +
                f"\n차트={has_chart} / 데이터프레임={has_df}\n\n"
                "다음 workflow 단계 4개 추천 (JSON 배열만):")
    try:
        raw = _call_ai_with_role([{"role": "user", "content": user_msg}], system=sys_prompt, role="analysis")
        m = re.search(r'\[.*?\]', raw, re.DOTALL)
        if m:
            parsed = json.loads(m.group())
            if isinstance(parsed, list) and all("text" in x for x in parsed):
                return [{"type": x.get("type", "analyze"), "text": x["text"]}
                        for x in parsed[:4]]
    except Exception:
        pass

    # Fallback: workflow 정적 추천
    seen: set[str] = set()
    result: list[dict] = []
    for step in plan.get("steps", []):
        for s in _WORKFLOW_FALLBACK.get(step.get("tool", ""), []):
            if s["text"] not in seen:
                seen.add(s["text"]); result.append(s)
            if len(result) >= 4:
                return result
    return result[:4]


# ══════════════════════════════════════════════
# 파일 로딩 — 병합 셀 해제 + 2행 헤더 자동 감지
# ══════════════════════════════════════════════

def _read_excel_smart(fp: Path) -> pd.DataFrame:
    """
    openpyxl 기반 엑셀 읽기.
    - 병합 셀을 좌상단 값으로 채워서 NaN 제거
    - 2행 헤더 자동 감지 (소제목이 대제목과 다르면 소제목 사용)
    - 쉼표 포함 숫자 문자열('51,840,000') → 숫자 변환
    """
    from itertools import zip_longest
    try:
        from openpyxl import load_workbook
        wb = load_workbook(fp, data_only=True)
        ws = wb.active

        # 헤더 영역(1~2행) 가로 병합의 보조 셀 수집 — 채우면 중복 컬럼명 발생
        _hdr_sec: set = set()
        for _mrg in ws.merged_cells.ranges:
            if _mrg.min_row <= 2 and _mrg.max_col > _mrg.min_col:
                for _r in range(_mrg.min_row, min(_mrg.max_row + 1, 3)):
                    for _c in range(_mrg.min_col + 1, _mrg.max_col + 1):
                        _hdr_sec.add((_r, _c))

        # 병합 셀 해제: 데이터 영역은 좌상단 값으로 채우기, 헤더 가로 보조셀은 제외
        for rng in list(ws.merged_cells.ranges):
            top = ws.cell(rng.min_row, rng.min_col).value
            ws.unmerge_cells(str(rng))
            for r in range(rng.min_row, rng.max_row + 1):
                for c in range(rng.min_col, rng.max_col + 1):
                    if (r, c) not in _hdr_sec:
                        ws.cell(r, c).value = top

        data = [[cell.value for cell in row] for row in ws.iter_rows()]
        wb.close()
        if not data:
            return pd.DataFrame()

        # 2행 헤더 감지 휴리스틱
        # 3번째 행에 숫자가 많고 2번째 행이 텍스트 → 2행 헤더 구조
        def _num_count(row):
            return sum(1 for v in row if isinstance(v, (int, float)) and v is not None)

        r0 = data[0] if len(data) > 0 else []
        r1 = data[1] if len(data) > 1 else []
        r2 = data[2] if len(data) > 2 else []
        two_row_header = (len(data) >= 3
                          and _num_count(r2) > _num_count(r1)
                          and _num_count(r1) < len([v for v in r1 if v is not None]) * 0.4)

        if two_row_header:
            raw: list = []
            prev_s1 = ""
            for h1, h2 in zip_longest(r0, r1, fillvalue=None):
                s1 = str(h1).strip() if h1 is not None else ""
                s2 = str(h2).strip() if h2 is not None else ""
                if not s1:
                    s1 = prev_s1  # 가로 병합 보조셀 — 부모 헤더 이어받기
                else:
                    prev_s1 = s1
                if s1 and s2 and s2 != s1:
                    raw.append(f"{s1}_{s2}")   # 부모_자식 (예: 실행예산_이월예산)
                elif s2:
                    raw.append(s2)
                elif s1:
                    raw.append(s1)
                else:
                    raw.append(f"col_{len(raw)}")
            data_rows = data[2:]
        else:
            prev_h = ""
            raw = []
            for i, v in enumerate(r0):
                h = str(v).strip() if v is not None else ""
                if not h:
                    h = prev_h  # 가로 병합 보조셀 — 부모 헤더 이어받기
                else:
                    prev_h = h
                raw.append(h if h else f"col_{i}")
            data_rows = data[1:]

        # 중복 컬럼명 처리
        seen: dict[str, int] = {}
        headers: list[str] = []
        for h in raw:
            if h in seen:
                seen[h] += 1; headers.append(f"{h}_{seen[h]}")
            else:
                seen[h] = 0; headers.append(h)

        if not data_rows:
            return pd.DataFrame(columns=headers)

        n_cols = max(len(headers), max(len(r) for r in data_rows))
        headers += [f"col_{i}" for i in range(len(headers), n_cols)]
        padded  = [r + [None] * (n_cols - len(r)) for r in data_rows]

        df = pd.DataFrame(padded, columns=headers[:n_cols])
        df = df.dropna(how="all").reset_index(drop=True)

        # 희소 텍스트 컬럼 ffill — 세로 병합 해제 후 남은 None 복원 (비목분류 등)
        for col in df.select_dtypes(exclude="number").columns:
            if df[col].notna().mean() < 0.5:
                df[col] = df[col].ffill()

        # 쉼표 포함 숫자 문자열 → 수치형 변환
        for col in df.columns:
            cleaned = df[col].apply(
                lambda x: str(x).replace(",", "") if isinstance(x, str) else x)
            conv = pd.to_numeric(cleaned, errors="coerce")
            if conv.notna().sum() / max(len(df), 1) >= 0.6:
                df[col] = conv

        return df
    except Exception:
        # openpyxl 실패 시 pandas 기본값
        return pd.read_excel(fp)


def load_files(filenames: list[str]) -> dict[str, pd.DataFrame]:
    dfs: dict[str, pd.DataFrame] = {}
    for fname in filenames:
        fp = UPLOAD_DIR / fname
        try:
            dfs[fname] = pd.read_csv(fp) if fp.suffix.lower() == ".csv" else _read_excel_smart(fp)
        except Exception:
            pass
    return dfs

# ══════════════════════════════════════════════
# 파일 구조 감사 + 검증 카드
# ══════════════════════════════════════════════

def _audit_file(fname: str, df: pd.DataFrame) -> dict:
    """파일 구조 감사 — 경고 및 통계 반환."""
    num_cols = df.select_dtypes(include="number").columns.tolist()
    cat_cols = df.select_dtypes(exclude="number").columns.tolist()
    sub_kws  = {"합계", "소계", "계", "총계", "합", "total", "sum", "subtotal"}
    sub_count = int(max(
        (df[c].astype(str).str.strip().str.lower().isin(sub_kws).sum() for c in cat_cols),
        default=0,
    ))
    null_ratio  = df.isnull().sum().sum() / max(df.size, 1)
    ambig_cols  = [c for c in df.columns if c.startswith("col_")]

    # 실제로 데이터가 채워진 행/열 범위
    filled_rows      = int(df.notna().any(axis=1).sum())          # 1개 이상 셀이 채워진 행
    text_filled_rows = int(df[cat_cols].notna().any(axis=1).sum()) if cat_cols else 0  # 텍스트 있는 행
    num_filled_rows  = int(df[num_cols].notna().any(axis=1).sum()) if num_cols else 0  # 숫자 있는 행
    text_col_count   = len(cat_cols)
    num_col_count    = len(num_cols)

    warnings: list[str] = []
    if sub_count:
        warnings.append(f"소계/합계 행 {sub_count}개 감지 → 집계 시 자동 제거됨")
    if ambig_cols:
        warnings.append(f"헤더 미확인 컬럼 {len(ambig_cols)}개 (col_N 형태 — 병합셀 원인)")
    if null_ratio > 0.15:
        warnings.append(f"빈 셀 {null_ratio*100:.0f}% 초과 — 데이터 정합성 확인 필요")
    return {
        "fname": fname, "rows": len(df), "cols": len(df.columns),
        "num_cols": num_cols[:6], "cat_cols": cat_cols[:5],
        "sub_count": sub_count, "warnings": warnings,
        "filled_rows": filled_rows,
        "text_filled_rows": text_filled_rows,
        "num_filled_rows": num_filled_rows,
        "text_col_count": text_col_count,
        "num_col_count": num_col_count,
    }


def _render_verification_card(tool: str, result: dict) -> None:
    """Tool 결과 아래 — 처리 기준·검증·샘플 근거 카드."""
    lines: list[str] = []

    if tool == "aggregate":
        gc   = result.get("group_by") or "전체"
        vc   = result.get("value_col") or "숫자 전체"
        fn   = result.get("func", "sum")
        df_r = result.get("dataframe")
        rows = len(df_r) if isinstance(df_r, pd.DataFrame) else 0
        lines += [
            f"<b>집계 기준</b> &nbsp; 그룹: <code>{gc}</code> &nbsp; 대상: <code>{vc}</code> &nbsp; 함수: <code>{fn}</code>",
            f"결과 <b>{rows}행</b> &nbsp;·&nbsp; 소계/합계 행 자동 제거 적용",
        ]
        per_file = result.get("per_file_sample", {})
        if per_file:
            rows_html = "".join(
                "<li><b>{k}</b>: {vals} → {fn} <b>{avg:,.0f}</b></li>".format(
                    k=k,
                    vals=" / ".join(f"{Path(fname).stem}: {v:,.0f}" for fname, v in fmap.items()),
                    fn=fn,
                    avg=sum(fmap.values()) / len(fmap),
                )
                for k, fmap in list(per_file.items())[:4]
            )
            lines.append(f"<b>샘플 근거 (파일별 원본값)</b><ul style='margin:4px 0 0 14px;padding:0'>{rows_html}</ul>")
        elif isinstance(df_r, pd.DataFrame) and not df_r.empty and gc in df_r.columns and vc in df_r.columns:
            top = df_r.nlargest(3, vc) if pd.api.types.is_numeric_dtype(df_r[vc]) else df_r.head(3)
            items = "".join(
                f"<li><b>{row.get(gc,'?')}</b>: {row.get(vc,0):,.0f}</li>"
                for _, row in top.iterrows()
                if pd.notna(row.get(vc))
            )
            if items:
                lines.append(f"<b>샘플 결과 (상위 3항목)</b><ul style='margin:4px 0 0 14px;padding:0'>{items}</ul>")

    elif tool == "compare_files":
        fa      = Path(result.get("file_a", "?")).stem
        fb      = Path(result.get("file_b", "?")).stem
        added   = len(result.get("added", []))
        removed = len(result.get("removed", []))
        changed = len(result.get("changes", []))
        lines += [
            f"<b>비교 대상</b> &nbsp; <code>{fa}</code> vs <code>{fb}</code>",
            f"추가 <b>{added}</b>건 &nbsp; 삭제 <b>{removed}</b>건 &nbsp; 변경 <b>{changed}</b>건",
        ]
        changes = result.get("changes", [])
        if changes:
            top3  = sorted(changes, key=lambda x: abs(x.get("증감", 0)), reverse=True)[:3]
            items = "".join(
                f"<li><b>{c.get('항목','?')}</b> [{c.get('컬럼','')}]: "
                f"{c.get('이전',0):,.0f} → {c.get('이후',0):,.0f} "
                f"({'+' if c.get('증감',0)>0 else ''}{c.get('증감',0):,.0f})</li>"
                for c in top3
            )
            lines.append(f"<b>주요 변경 (증감 상위 3)</b><ul style='margin:4px 0 0 14px;padding:0'>{items}</ul>")

    elif tool == "calculate_ratio":
        pc     = result.get("plan_col", "?")
        ec     = result.get("exec_col", "?")
        total_r = result.get("total_ratio", 0)
        tp, te  = result.get("total_plan", 0), result.get("total_exec", 0)
        lines += [
            f"<b>계획 컬럼</b>: <code>{pc}</code> &nbsp; <b>실행 컬럼</b>: <code>{ec}</code>",
            f"전체 집행률 <b>{total_r}%</b> &nbsp; (계획 {tp:,.0f} / 실행 {te:,.0f})",
        ]
        low = [r for r in result.get("rows", []) if r.get("집행률(%)") is not None and r["집행률(%)"] < 30][:3]
        if low:
            items = "".join(f"<li><b>{r.get('항목','?')}</b>: {r.get('집행률(%)','?')}%</li>" for r in low)
            lines.append(f"<b>⚠ 집행률 30% 미만 항목</b><ul style='margin:4px 0 0 14px;padding:0'>{items}</ul>")

    elif tool == "clean_table":
        orig  = result.get("original_rows", 0)
        after = result.get("cleaned_rows", 0)
        rm    = result.get("removed_rows", 0)
        lines += [
            f"원본 <b>{orig}행</b> → 정리 후 <b>{after}행</b>",
            f"제거된 행: <b>{rm}개</b> (소계·빈행 포함)",
        ]

    elif tool == "export":
        lines += [
            f"저장 파일: <b>{result.get('filename','?')}</b>",
            f"크기: <b>{result.get('rows',0)}행 × {result.get('cols',0)}열</b>",
        ]

    else:
        return

    body = "<br>".join(lines)
    st.markdown(
        f'<div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;'
        f'padding:10px 14px;margin:6px 0 2px;font-size:12.5px;color:#374151;line-height:1.75">'
        f'{body}</div>',
        unsafe_allow_html=True,
    )


def get_file_summaries(filenames: list[str], dfs: dict) -> str:
    parts = []
    for fname in filenames:
        df = dfs.get(fname)
        if df is None: continue
        cols = ", ".join(
            f"{_excel_col_name(i)}={col}({'N' if pd.api.types.is_numeric_dtype(df[col]) else 'T'})"
            for i, col in enumerate(df.columns[:15])
        )
        sample = df.head(5).to_string(index=False, max_cols=12)
        parts.append(
            f"- {fname}: {df.shape[0]}행×{df.shape[1]}열 | {cols}\n"
            f"  샘플(처음 5행):\n{sample}"
        )
    return "\n".join(parts)

def build_full_context(filenames: list[str], dfs: dict) -> str:
    """Fallback용 풍부한 컨텍스트 (실제 데이터 포함)."""
    parts = []
    for fname in filenames:
        df = dfs.get(fname)
        if df is None: continue
        col_lines = []
        for i, col in enumerate(df.columns):
            ecol = _excel_col_name(i)
            if pd.api.types.is_numeric_dtype(df[col]):
                col_lines.append(f"  {ecol}열[{col}]: 숫자(합={df[col].sum():,.0f}/평균={df[col].mean():,.1f})")
            else:
                uniq = df[col].dropna().unique()
                col_lines.append(f"  {ecol}열[{col}]: 텍스트({df[col].dropna().nunique()}개 — 예:{list(uniq[:5])})")
        n   = min(len(df), 30)
        dfs2 = df.head(n).copy(); dfs2.index = range(2, n + 2)
        parts.append(
            f"### {fname}\n컬럼:\n" + "\n".join(col_lines) +
            f"\n\n실제 데이터(Excel 행 번호):\n```\n{dfs2.to_string()}\n```\n"
        )
    return "\n".join(parts) or "선택된 파일 없음"

def detect_target_files(question: str, filenames: list[str]) -> list[str]:
    q = question.lower()
    mentioned = [f for f in filenames if f.lower() in q or Path(f).stem.lower() in q]
    return mentioned if mentioned else filenames


# ══════════════════════════════════════════════
# AI 호출
# ══════════════════════════════════════════════

def _route_from_session(prompt: str = "", role: str | None = None) -> RouteResult | None:
    """세션 또는 프롬프트/역할로 라우트 결정."""
    if role:
        return route_for_role(role, prompt, load_model_config())
    route = st.session_state.get("auto_route")
    if route and (not prompt or route.task_type == classify_task(prompt)):
        return route
    if prompt:
        return route_for_prompt(prompt, load_model_config())
    return route_for_role("analysis", "", load_model_config())


def call_ai_raw(
    messages: list,
    system: str = "",
    *,
    prompt: str = "",
    role: str | None = None,
    task_type: str | None = None,
) -> str:
    """작업 유형별 자동 모델로 LLM 호출."""
    cfg = load_model_config()
    if not has_model_pool(cfg):
        return "⚠️ 모델 풀이 비어 있습니다. Model Manager에서 모델을 연결하세요."

    if task_type:
        from utils.task_router import resolve_route
        route = resolve_route(task_type, cfg)
    else:
        user_text = prompt or next(
            (m["content"] for m in reversed(messages) if m.get("role") == "user"), ""
        )
        route = _route_from_session(user_text, role)

    if route is None:
        return "⚠️ 사용 가능한 모델이 없습니다."

    st.session_state["auto_route"] = route
    out = call_with_route(messages, route, system=system)
    if route.provider == "openai" and out and not out.startswith("⚠️"):
        try:
            import openai
            # 토큰은 call_with_route 내부에서 집계 어려워 추정 생략
            pass
        except Exception:
            pass
    return out


def stream_ai_raw(
    messages: list,
    system: str = "",
    *,
    prompt: str = "",
    role: str | None = None,
    task_type: str | None = None,
):
    """작업 유형별 자동 모델 스트리밍."""
    if task_type:
        from utils.task_router import resolve_route
        route = resolve_route(task_type, load_model_config())
    else:
        user_text = prompt or next(
            (m["content"] for m in reversed(messages) if m.get("role") == "user"), ""
        )
        route = _route_from_session(user_text, role)
    if route is None:
        yield "⚠️ 사용 가능한 모델이 없습니다."
        return
    st.session_state["auto_route"] = route
    yield from stream_with_route(messages, route, system=system)


def _call_ai_with_role(messages: list, system: str = "", role: str = "analysis") -> str:
    """역할 → 자동 모델 선택 후 호출."""
    user_text = next((m["content"] for m in reversed(messages) if m.get("role") == "user"), "")
    return call_ai_raw(messages, system, prompt=user_text, role=role)


def _stream_ai_with_role(messages: list, system: str = "", role: str = "analysis"):
    """역할 → 자동 모델 스트리밍."""
    user_text = next((m["content"] for m in reversed(messages) if m.get("role") == "user"), "")
    yield from stream_ai_raw(messages, system, prompt=user_text, role=role)


def call_planner(question: str, file_summaries: str) -> dict | None:
    """Phase 1: 실행 계획 JSON 생성."""
    sys_prompt = f"""너는 엑셀 작업 계획자야. 사용자 질문을 분석해서 실행 계획을 JSON으로만 출력해.
JSON 외 다른 텍스트 절대 없이. 반드시 valid JSON만.

사용 가능한 도구:
{TOOL_DESCRIPTIONS}

파일 목록:
{file_summaries}

출력 형식:
{{"summary":"한 줄 설명","steps":[{{"tool":"도구명","args":{{"key":"value"}}}}]}}"""
    st.session_state["auto_route"] = route_for_prompt(question, load_model_config())
    raw = _call_ai_with_role([{"role": "user", "content": question}], system=sys_prompt, role="analysis")
    try:
        m = re.search(r'\{.*\}', raw, re.DOTALL)
        if m: return json.loads(m.group())
    except (json.JSONDecodeError, AttributeError):
        pass
    return None


def call_explainer(question: str, plan: dict, results: list) -> str:
    """Phase 3: 실제 결과를 받아 한국어로 설명."""
    results_text = []
    for item in results:
        tool, res = item["tool"], item["result"]
        display = {k: v for k, v in res.items() if k not in ("dataframe", "chart_b64")}
        results_text.append(f"[{tool}] 결과:\n{json.dumps(display, ensure_ascii=False, indent=2, default=str)}")
    sys_prompt = """너는 엑셀 데이터 분석 전문가야. Python 도구가 실제로 계산한 결과를 한국어로 설명해.

규칙:
1. "샘플", "추정", "아마" 금지 — 실제 계산 결과만 인용
2. 셀 위치 포함 (예: "B3=121", "C열: 내부인건비")
3. 답변 구조: ① 한 줄 요약 → ② 실제 결과 → ③ 다음 작업 제안 2-3개
4. 한국어로만"""
    prompt = f"""사용자 질문: {question}
실행 계획: {plan.get('summary', '')}

실제 분석 결과:
{chr(10).join(results_text)}"""
    return call_ai_raw([{"role": "user", "content": prompt}], system=sys_prompt, prompt=question, role=role)


def stream_explainer(question: str, plan: dict, results: list, role: str = "analysis"):
    """Phase 3: 설명 스트리밍. role은 plan의 첫 번째 tool에서 결정."""
    results_text = []
    for item in results:
        tool, res = item["tool"], item["result"]
        display = {k: v for k, v in res.items() if k not in ("dataframe", "chart_b64")}
        results_text.append(f"[{tool}] 결과:\n{json.dumps(display, ensure_ascii=False, indent=2, default=str)}")
    sys_prompt = """너는 엑셀 데이터 분석 전문가야. Python 도구가 실제로 계산한 결과를 한국어로 설명해.

규칙:
1. "샘플", "추정", "아마" 금지 — 실제 계산 결과만 인용
2. 셀 위치 포함 (예: "B3=121", "C열: 내부인건비")
3. 답변 구조: ① 한 줄 요약 → ② 실제 결과 → ③ 다음 작업 제안 2-3개
4. 한국어로만"""
    prompt = f"""사용자 질문: {question}
실행 계획: {plan.get('summary', '')}

실제 분석 결과:
{chr(10).join(results_text)}"""
    yield from _stream_ai_with_role([{"role": "user", "content": prompt}], system=sys_prompt, role=role)


def call_ai_fallback(messages: list, file_context: str) -> str:
    """Fallback: 코드 생성 방식."""
    system = textwrap.dedent(f"""\
        당신은 한국 행정 담당자를 위한 엑셀 분석 전문가입니다.
        파일의 실제 데이터가 제공됩니다 (샘플 아님).

        규칙:
        1. "샘플 데이터 기준" 표현 금지 — 실제 데이터입니다.
        2. 데이터 값 인용 시 셀 위치 포함 (B3=121, C열 등)
        3. 파일 조작이 필요하면 ```python 블록으로 코드 생성
           - UPLOAD_DIR, OUTPUT_DIR 변수 사용 (이미 정의됨)
           - 출력 파일명에 타임스탬프 포함
           - print()로 한국어 결과 요약
        4. 단순 조회는 코드 없이 직접 답변
        5. 한국어로 답변

        파일 데이터:
        {file_context}
    """)
    return call_ai_raw(messages, system=system)


def stream_ai_fallback(messages: list, file_context: str):
    """Fallback: 코드 생성 방식 (스트리밍)."""
    system = textwrap.dedent(f"""\
        당신은 한국 행정 담당자를 위한 엑셀 분석 전문가입니다.
        파일의 실제 데이터가 제공됩니다 (샘플 아님).

        ## 실행 환경 (exec 컨텍스트)
        - `files` : dict[str, DataFrame] — 선택된 파일이 이미 로드됨
          예) `for fname, df in files.items():`  또는  `df = files["파일명.xlsx"]`
        - `save(df, "결과.xlsx")` : DataFrame을 outputs 폴더에 저장하는 함수
        - `result = df` : 이 변수에 DataFrame을 저장하면 자동으로 표로 표시됨
        - `pd`, `plt`, `io`, `Path`, `datetime` 사전 로드됨
        - import 문 사용 금지 (pd, plt 이미 제공됨)

        ## 규칙
        1. "샘플 데이터 기준" 표현 금지 — 실제 데이터입니다.
        2. 데이터 값 인용 시 셀 위치 포함 (B3=121, C열 등)
        3. 파일 조작이 필요하면 ```python 블록으로 코드 생성
           - 파일 접근: `files["파일명.xlsx"]` 사용 (pd.read_excel 불필요)
           - 결과 표시: `result = 최종_DataFrame`
           - 파일 저장: `save(result, "결과파일명.xlsx")`
           - print()로 한국어 결과 요약 추가
        4. 단순 조회는 코드 없이 직접 답변
        5. 한국어로 답변

        ## 파일 데이터
        {file_context}
    """)
    yield from stream_ai_raw(messages, system=system)


# ══════════════════════════════════════════════
# 코드 실행 (Fallback)
# ══════════════════════════════════════════════

def extract_and_run_code(response_text: str):
    import builtins as _bm, signal as _sig

    blocks = re.findall(r"```python\s*\n(.*?)```", response_text, re.DOTALL)
    if not blocks:
        return None, None, None, [], [], None
    code = blocks[0]
    plt.close("all")

    # ── pre-load DataFrames → files dict (SheetPilot 방식) ──
    _exec_dfs: dict[str, pd.DataFrame] = {}
    for _fn in st.session_state.get("selected_files", []):
        _fp = UPLOAD_DIR / _fn
        if _fp.exists():
            try:
                _exec_dfs[_fn] = (pd.read_csv(_fp) if _fp.suffix.lower() == ".csv"
                                  else _read_excel_smart(_fp))
            except Exception:
                pass

    # ── save() 헬퍼 ──
    def _save_helper(df: pd.DataFrame, filename: str = "result.xlsx") -> str:
        if not isinstance(df, pd.DataFrame):
            raise ValueError("save()의 첫 번째 인자는 DataFrame이어야 합니다.")
        safe = Path(filename).name
        if not safe.endswith((".xlsx", ".csv")):
            safe += ".xlsx"
        out = OUTPUT_DIR / safe
        if safe.endswith(".csv"):
            df.to_csv(out, index=False, encoding="utf-8-sig")
        else:
            df.to_excel(out, index=False)
        add_entry("AI 처리", safe, "executed", "save() via code")
        return str(out)

    # ── 제한된 builtins (open, eval, exec 등 차단) ──
    _BLOCKED = {"open", "input", "eval", "exec", "compile", "__import__",
                "getattr", "setattr", "delattr", "vars", "globals", "locals"}
    _safe_builtins = {k: v for k, v in vars(_bm).items() if k not in _BLOCKED}

    exec_g = {
        "__builtins__": _safe_builtins,
        "pd": pd, "Path": Path, "datetime": datetime,
        "UPLOAD_DIR": UPLOAD_DIR, "OUTPUT_DIR": OUTPUT_DIR,
        "plt": plt, "matplotlib": matplotlib, "io": io,
        "files": _exec_dfs,   # pre-loaded DataFrames
        "save":  _save_helper, # save(df, "output.xlsx")
        "result": None,        # LLM이 결과를 여기에 저장
    }

    stdout_buf = io.StringIO()
    old = sys.stdout; sys.stdout = stdout_buf
    error = None
    out_before = set(OUTPUT_DIR.glob("*"))

    # ── 30초 타임아웃 (POSIX) ──
    def _timeout_handler(signum, frame):
        raise TimeoutError("코드 실행이 30초를 초과했습니다 (무한루프 또는 과도한 연산)")
    _sig.signal(_sig.SIGALRM, _timeout_handler)
    _sig.alarm(30)
    try:
        exec(code, exec_g)  # noqa: S102
    except TimeoutError as e:
        error = str(e)
    except Exception:
        error = traceback.format_exc()
    finally:
        _sig.alarm(0)
        sys.stdout = old

    stdout_out = stdout_buf.getvalue()
    result_df  = exec_g.get("result")
    if not isinstance(result_df, pd.DataFrame):
        result_df = None

    charts = []
    for n in plt.get_fignums():
        fig = plt.figure(n); buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=120, bbox_inches="tight"); buf.seek(0)
        charts.append(base64.b64encode(buf.read()).decode())
    plt.close("all")
    new_files = [str(f) for f in set(OUTPUT_DIR.glob("*")) - out_before]
    return code, stdout_out, error, charts, new_files, result_df


# ══════════════════════════════════════════════
# 코드 안전성 분석 (AST)
# ══════════════════════════════════════════════

def _check_code_safety(code: str) -> tuple[bool, list[str]]:
    """AST walk으로 위험 패턴 감지. (is_safe, warnings) 반환."""
    import ast as _ast
    _BLOCKED_MODS  = {"os", "subprocess", "shutil", "socket", "paramiko", "pty", "ctypes"}
    _BLOCKED_CALLS = {"eval", "exec", "compile", "__import__"}
    _BLOCKED_ATTRS = {
        "system", "popen", "run", "call", "Popen", "check_output",
        "rmtree", "remove", "unlink", "chmod", "chown",
    }
    try:
        tree = _ast.parse(code)
    except SyntaxError as e:
        return False, [f"문법 오류: {e}"]

    warns: list[str] = []
    for node in _ast.walk(tree):
        if isinstance(node, (_ast.Import, _ast.ImportFrom)):
            names = ([n.name for n in node.names] if isinstance(node, _ast.Import)
                     else [node.module or ""])
            for nm in names:
                if nm.split(".")[0] in _BLOCKED_MODS:
                    warns.append(f"위험 모듈 import: `{nm.split('.')[0]}`")
        if isinstance(node, _ast.Call):
            func = node.func
            if isinstance(func, _ast.Name) and func.id in _BLOCKED_CALLS:
                warns.append(f"위험 함수: `{func.id}()`")
            if isinstance(func, _ast.Attribute) and func.attr in _BLOCKED_ATTRS:
                warns.append(f"위험 메서드: `.{func.attr}()`")

    seen: set[str] = set()
    deduped = [w for w in warns if not (w in seen or seen.add(w))]  # type: ignore[func-returns-value]
    return len(deduped) == 0, deduped


# ══════════════════════════════════════════════
# 코드 실행 승인 UI
# ══════════════════════════════════════════════

def _render_code_approval() -> None:
    """마지막 메시지의 pending 코드에 대한 실행 승인 UI."""
    msgs = st.session_state.messages
    if not msgs:
        return
    last   = msgs[-1]
    code   = last.get("_pending_code", "")
    warns  = last.get("_code_warns", [])
    prompt = st.session_state.get("_exec_prompt", "")

    badge = (
        '<span style="background:#d1fae5;color:#065f46;padding:3px 10px;'
        'border-radius:8px;font-size:12px;font-weight:600">🟢 안전 — 위험 패턴 없음</span>'
        if not warns else
        f'<span style="background:#fef3c7;color:#92400e;padding:3px 10px;'
        f'border-radius:8px;font-size:12px;font-weight:600">⚠ 주의 — {len(warns)}개 패턴 감지</span>'
    )
    warn_items = "".join(
        f'<li style="font-size:12px;color:#92400e;margin-top:3px">{w}</li>' for w in warns
    )
    st.markdown(
        f'<div style="border:1px solid #e2e8f0;border-radius:12px;padding:14px 18px;'
        f'background:#f8fafc;margin:10px 0 6px">'
        f'<div style="font-size:13px;font-weight:700;color:#334155;margin-bottom:8px">'
        f'🔬 코드를 실행할까요?</div>'
        f'{badge}'
        + (f'<ul style="margin:6px 0 0 16px;padding:0">{warn_items}</ul>' if warns else "")
        + '</div>',
        unsafe_allow_html=True,
    )

    btn_lbl = "그래도 실행" if warns else "실행하기 ▶"
    col_run, col_cancel, _ = st.columns([1.3, 1, 4])

    if col_run.button(btn_lbl, type="primary", use_container_width=True, key="approve_run"):
        wrapped = f"```python\n{code}\n```"
        _c, _out, _err, _charts, _files, _rdf = extract_and_run_code(wrapped)
        last.pop("_pending_code", None)
        last.pop("_code_safe", None)
        last.pop("_code_warns", None)
        st.session_state.pop("_exec_prompt", None)
        if _out:                         last["code_output"]  = _out
        if _err:                         last["code_error"]   = _err
        if _charts:                      last["charts"]       = _charts
        if _files:                       last["output_files"] = _files
        if isinstance(_rdf, pd.DataFrame): last["result_df"]  = _rdf
        if _c and not _err:
            save_to_history(prompt, _c)
            add_entry("AI 처리", Path(_files[0]).name if _files else "result",
                      "executed", f"approved: {prompt[:60]}")
        st.rerun()

    if col_cancel.button("취소", use_container_width=True, key="approve_cancel"):
        last.pop("_pending_code", None)
        last.pop("_code_safe", None)
        last.pop("_code_warns", None)
        st.session_state.pop("_exec_prompt", None)
        st.rerun()


# ══════════════════════════════════════════════
# 메시지 렌더링
# ══════════════════════════════════════════════

_SUGG_CSS = """
<style>
[data-testid="stColumn"]:has(.sugg-marker) button {
    border: 1px solid #e2e8f0 !important;
    font-size: 12px !important;
    text-align: left !important;
    height: auto !important;
    min-height: 54px !important;
    white-space: normal !important;
    line-height: 1.45 !important;
    padding: 10px 12px !important;
    border-radius: 8px !important;
    transition: filter .15s !important;
}
[data-testid="stColumn"]:has(.sugg-marker) button:hover {
    filter: brightness(0.93) !important;
}
[data-testid="stColumn"]:has(.sugg-marker[data-type="analyze"]) button {
    background: #eff6ff !important; border-color: #bfdbfe !important; color: #1e40af !important;
}
[data-testid="stColumn"]:has(.sugg-marker[data-type="clean"]) button {
    background: #fffbeb !important; border-color: #fde68a !important; color: #92400e !important;
}
[data-testid="stColumn"]:has(.sugg-marker[data-type="save"]) button {
    background: #f0fdf4 !important; border-color: #bbf7d0 !important; color: #065f46 !important;
}
[data-testid="stColumn"]:has(.sugg-marker[data-type="insight"]) button {
    background: #faf5ff !important; border-color: #e9d5ff !important; color: #5b21b6 !important;
}
</style>
"""

def render_message(msg: dict, sections: set[str] | None = None):
    """sections: None=전체, 또는 plan/content/tools/fallback/suggestions."""
    def _show(name: str) -> bool:
        return sections is None or name in sections

    if _show("plan") and msg.get("plan"):
        plan = msg["plan"]
        with st.expander(f"실행 계획 — {plan.get('summary', '')}", expanded=False):
            for i, s in enumerate(plan.get("steps", []), 1):
                st.markdown(f"**{i}.** `{s['tool']}` &nbsp; `{json.dumps(s.get('args',{}), ensure_ascii=False)}`")

    if _show("content"):
        content    = msg.get("content", "")
        code_bks   = re.findall(r"```python\s*\n(.*?)```", content, re.DOTALL)
        clean_text = re.sub(r"```python\s*\n.*?```", "", content, flags=re.DOTALL).strip()
        if clean_text and not msg.get("_streamed"):
            st.markdown(clean_text)
        for i, cb in enumerate(code_bks):
            lbl = "Python 코드 보기" if len(code_bks) == 1 else f"Python 코드 {i+1} 보기"
            with st.expander(lbl, expanded=False):
                st.code(cb, language="python")

    if _show("tools"):
        for item in msg.get("tool_results", []):
            res = item["result"]
            if "chart_b64" in res:
                st.image(base64.b64decode(res["chart_b64"]), use_container_width=True)
            if isinstance(res.get("dataframe"), pd.DataFrame) and not res["dataframe"].empty:
                df_show = res["dataframe"]
                label   = f"{item['tool']} 결과 ({len(df_show):,}행)"
                with st.expander(label, expanded=True):
                    st.dataframe(df_show.head(50).map(
                        lambda x: f"{x:,.1f}" if isinstance(x, float) and not pd.isna(x) else x
                    ), use_container_width=True)
                    if len(df_show) > 50:
                        st.caption(f"처음 50행 / 전체 {len(df_show):,}행")
            # 처리 기준·검증·샘플 근거 카드
            _render_verification_card(item["tool"], res)
            if "output_file" in res:
                fp = Path(res["output_file"])
                if fp.exists():
                    with open(fp, "rb") as fh:
                        st.download_button(f"⬇ {fp.name} 다운로드", data=fh.read(),
                                           file_name=fp.name, key=f"tdl_{fp.name}_{id(msg)}")

    if _show("fallback"):
        # result 변수 DataFrame 자동 표시 (SheetPilot 방식)
        if isinstance(msg.get("result_df"), pd.DataFrame):
            _rdf = msg["result_df"]
            with st.expander(f"실행 결과 ({len(_rdf):,}행)", expanded=True):
                st.dataframe(_rdf.head(50).map(
                    lambda x: f"{x:,.1f}" if isinstance(x, float) and not pd.isna(x) else x
                ), use_container_width=True)
                if len(_rdf) > 50:
                    st.caption(f"처음 50행 / 전체 {len(_rdf):,}행")
            st.success("Code executed successfully.")
        if msg.get("code_output", "").strip():
            with st.expander("출력 (print)", expanded=not msg.get("result_df")):
                st.code(msg["code_output"], language="text")
        if "code_error" in msg:
            with st.expander("실행 오류", expanded=True):
                st.code(msg["code_error"], language="text")
        for c in msg.get("charts", []):
            st.image(base64.b64decode(c), use_container_width=True)
        for of in msg.get("output_files", []):
            fp = Path(of)
            if fp.exists():
                with open(fp, "rb") as fh:
                    st.download_button(f"⬇ {fp.name} 다운로드", data=fh.read(),
                                       file_name=fp.name, key=f"fdl_{fp.name}_{id(msg)}")

    if _show("suggestions"):
        suggs = msg.get("suggestions", [])
        if suggs:
            st.markdown(_SUGG_CSS, unsafe_allow_html=True)
            st.markdown(
                '<div style="margin-top:14px;font-size:11px;color:#94a3b8;'
                'font-weight:600;letter-spacing:.04em;margin-bottom:6px">'
                '다음 작업 추천</div>',
                unsafe_allow_html=True,
            )
            uid  = msg.get("_id", str(id(msg)))
            n    = min(len(suggs), 4)
            # 4개 이하면 2×2, 2개 이하면 1행
            grid_cols = st.columns(2) if n > 2 else st.columns(n)
            for ci, sugg in enumerate(suggs[:4]):
                if isinstance(sugg, dict):
                    stype = sugg.get("type", "analyze")
                    text  = sugg.get("text", "")
                else:
                    stype, text = "analyze", sugg
                icon = _TYPE_META.get(stype, ("📊",))[0]
                col  = grid_cols[ci % len(grid_cols)]
                col.markdown(
                    f'<span class="sugg-marker" data-type="{stype}" style="display:none"></span>',
                    unsafe_allow_html=True,
                )
                if col.button(f"{icon}  {text}", key=f"sugg_{uid}_{ci}",
                              use_container_width=True):
                    st.session_state["_pending"] = text
                    st.rerun()


# ══════════════════════════════════════════════
# 메인 UI
# ══════════════════════════════════════════════

_h_title, _h_btn = st.columns([5, 1])
_h_title.markdown("## AI Prompt")
with _h_btn:
    st.markdown("<div style='padding-top:12px'>", unsafe_allow_html=True)
    if st.button("대화 초기화", use_container_width=True, help="채팅 기록을 모두 지웁니다"):
        st.session_state["messages"] = []
        st.session_state["tokens_used"] = 0
        st.session_state.pop("_pending", None)
        st.session_state.pop("_exec_prompt", None)
        st.rerun()
    st.markdown("</div>", unsafe_allow_html=True)

# ── 선택 파일 태그 ──
if selected_files:
    tags = "".join(
        f'<span style="background:#dbeafe;color:#1e40af;padding:3px 10px;'
        f'border-radius:20px;font-size:12px;font-weight:500;margin-right:6px">{fn}</span>'
        for fn in selected_files[:6]
    )
    if len(selected_files) > 6:
        tags += f'<span style="color:#64748b;font-size:12px">+{len(selected_files)-6}개 더</span>'
    st.markdown(tags, unsafe_allow_html=True)

# ── 빠른 명령 버튼 ──
if selected_files:
    st.markdown('<div style="font-size:11px;color:#94a3b8;margin:8px 0 4px">빠른 명령</div>',
                unsafe_allow_html=True)
    _PRESETS = [
        ("열 설명",    "각 열이 무슨 의미인지 설명해줘"),
        ("집행률 계산", "계획 대비 집행률을 계산해줘"),
        ("이상값 찾기", "비어있거나 0인 값, 이상한 패턴을 찾아줘"),
        ("소계 제외",  "합계·소계 행을 제외하고 데이터만 정리해줘"),
        ("파일 비교",  "선택된 파일들을 비교해줘"),
        ("데이터 추출", "데이터를 전체 추출해서 보여줘"),
        ("엑셀 저장",  "정리된 데이터를 새 엑셀 파일로 저장해줘"),
        ("차트 생성",  "주요 수치를 막대 차트로 시각화해줘"),
    ]
    for row_start in range(0, len(_PRESETS), 4):
        row_items = _PRESETS[row_start:row_start + 4]
        cols = st.columns(4)
        for i, (lbl, cmd) in enumerate(row_items):
            if cols[i].button(lbl, key=f"preset_{row_start+i}", use_container_width=True):
                st.session_state["_pending"] = cmd; st.rerun()

st.markdown("")

# ── 히스토리 재실행 ──
if "rerun_prompt" in st.session_state:
    _rp = st.session_state.pop("rerun_prompt")
    _rc = st.session_state.pop("rerun_code", None)
    st.session_state.messages.append({"role": "user", "content": f"[재실행] {_rp}"})
    if _rc:
        code, stdout_out, error, charts, new_files, result_df = extract_and_run_code(f"```python\n{_rc}\n```")
        _rd = {"role": "assistant", "content": f"코드를 재실행했습니다.\n\n```python\n{_rc}\n```"}
        if stdout_out:                        _rd["code_output"] = stdout_out
        if error:                             _rd["code_error"]  = error
        if charts:                            _rd["charts"]      = charts
        if new_files:                         _rd["output_files"]= new_files
        if isinstance(result_df, pd.DataFrame): _rd["result_df"] = result_df
        st.session_state.messages.append(_rd)

# ── 메시지 히스토리 ──
for msg in st.session_state.messages:
    with st.chat_message(msg["role"]):
        render_message(msg)

_last_msg = st.session_state.messages[-1] if st.session_state.messages else {}
if _last_msg.get("_pending_code"):
    _render_code_approval()
    prompt = None
else:
    _pending = st.session_state.pop("_pending", None)
    typed    = st.chat_input("파일에 대해 무엇이든 물어보세요  (예: 'B열 값 나열', '집행률 계산', '두 파일 비교')")
    prompt   = typed or _pending

if prompt:
    st.session_state.messages.append({"role": "user", "content": prompt})
    with st.chat_message("user"):
        st.markdown(prompt)

    with st.chat_message("assistant"):
        target_files = detect_target_files(prompt, selected_files) if selected_files else []
        msg_data: dict = {"role": "assistant", "content": "",
                          "_id": datetime.now().strftime("%Y%m%d%H%M%S%f")}

        with st.status("데이터를 분석하는 중…", expanded=True) as proc_status:

            # 파일 로드
            if not target_files:
                st.write("선택된 파일이 없습니다.")
                dfs = {}
            else:
                if len(target_files) < len(selected_files):
                    st.write(f"질문에서 파일 감지: **{', '.join(target_files)}**")
                else:
                    st.write(f"파일 {len(target_files)}개 로드 중…")
                dfs = load_files(target_files)
                for fname, df in dfs.items():
                    audit = _audit_file(fname, df)
                    cat_p = " / ".join(audit["cat_cols"][:3])
                    num_p = " / ".join(audit["num_cols"][:3])
                    hdr_p = " / ".join(filter(None, [cat_p, num_p]))
                    st.write(f"  ✅ **{fname}** — {audit['rows']}행 × {audit['cols']}열")
                    if hdr_p:
                        st.write(f"     헤더: {hdr_p} ···")
                    for w in audit["warnings"]:
                        st.write(f"     ⚠ {w}")

            # ── Phase 1: 계획 생성 ──
            if dfs:
                st.write("실행 계획 생성 중…")
                summaries = get_file_summaries(target_files, dfs)
                plan      = call_planner(prompt, summaries)
            else:
                # 파일 미선택이어도 UPLOAD_DIR에 파일이 있으면 Planner 호출
                # (list_files 같은 도구는 dfs 없이 디렉터리 직접 스캔)
                _all_uploads = sorted(
                    [f for f in UPLOAD_DIR.glob("*")
                     if f.suffix.lower() in {".xlsx", ".xls", ".csv"} and f.name != ".gitkeep"],
                    key=lambda f: f.stat().st_mtime, reverse=True,
                )
                if _all_uploads:
                    _ul_list = "\n".join(
                        f"- {f.name}  ({f.stat().st_size // 1024} KB)"
                        for f in _all_uploads
                    )
                    summaries = f"현재 선택된 파일 없음. 업로드된 파일 목록:\n{_ul_list}"
                    plan = call_planner(prompt, summaries)
                else:
                    plan = None

            if plan and plan.get("steps"):
                st.write(f"계획: **{plan.get('summary', '')}**")
                for i, s in enumerate(plan["steps"], 1):
                    st.write(f"  {i}. `{s['tool']}`")

                # ── Phase 2: Tool 실행 ──
                ctx: dict = {}
                tool_results: list = []

                for step in plan["steps"]:
                    tname = step.get("tool", "")
                    targs = step.get("args", {})
                    fn    = TOOL_REGISTRY.get(tname)
                    if not fn:
                        st.write(f"  알 수 없는 도구: {tname}")
                        continue
                    st.write(f"  {tname} 실행 중…")
                    try:
                        result = fn(targs, ctx, dfs)
                    except Exception as e:
                        result = {"text": f"오류: {e}"}
                    tool_results.append({"tool": tname, "result": result})
                    ctx[tname] = result
                    if "output_name" in result:
                        ctx[result["output_name"]] = result
                    if "output_file" in result:
                        add_entry("AI 처리", Path(result["output_file"]).name, "executed",
                                  f"tool:{tname}")

                proc_status.update(label="도구 실행 완료", state="complete", expanded=False)

                msg_data["plan"]         = plan
                msg_data["tool_results"] = tool_results
                render_message(msg_data, sections={"plan", "tools"})

                # ── Phase 3: 설명 (스트리밍) ──
                _plan_role = TOOL_ROLES.get(
                    plan.get("steps", [{}])[0].get("tool", ""), "analysis"
                )
                explanation = st.write_stream(stream_explainer(prompt, plan, tool_results, role=_plan_role))
                msg_data["content"]  = explanation
                msg_data["_streamed"] = True

                # ── 추천: 설명 완료 후 실제 결과 기반으로 LLM 생성 ──
                msg_data["suggestions"] = generate_suggestions_ai(prompt, plan, tool_results)
                render_message(msg_data, sections={"suggestions"})

            else:
                # ── Fallback: 코드 생성 방식 ──
                if plan is None:
                    st.write("계획 생성 실패 — 코드 생성 방식으로 처리 중…")
                else:
                    st.write("코드 생성 방식으로 처리 중…")
                ctx_str   = build_full_context(target_files, dfs)
                hist      = [{"role": m["role"], "content": m["content"]}
                             for m in st.session_state.messages]
                proc_status.update(label="처리 완료", state="complete", expanded=False)
                st.session_state["auto_route"] = route_for_role("code", prompt, load_model_config())
                ai_resp = st.write_stream(stream_ai_fallback(hist, ctx_str))
                msg_data["content"] = ai_resp
                msg_data["_streamed"] = True
                _fb_blocks = re.findall(r"```python\s*\n(.*?)```", ai_resp, re.DOTALL)
                if _fb_blocks:
                    _fb_code = _fb_blocks[0]
                    _fb_safe, _fb_warns = _check_code_safety(_fb_code)
                    msg_data["_pending_code"] = _fb_code
                    msg_data["_code_safe"]    = _fb_safe
                    msg_data["_code_warns"]   = _fb_warns
                    st.session_state["_exec_prompt"] = prompt

        if not msg_data.get("_streamed"):
            render_message(msg_data)
        else:
            render_message(msg_data, sections={"content", "fallback"})
        msg_data.pop("_streamed", None)

    st.session_state.messages.append(msg_data)
    if msg_data.get("_pending_code"):
        st.rerun()
